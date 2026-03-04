"""VIES VAT Validator (Slango-style)

Aplicación de escritorio para validar números VAT (VIES) a partir de un Excel.
- Carga Excel, detecta columna VAT/NIF automáticamente.
- Separa resultados en pestañas: Pendientes / Validados.
- Maneja limitaciones del servicio (THROTTLED / TIMEOUT) sin bucles infinitos.
- Acción rápida "Abrir VIES": copia el número sin prefijo y abre la web.

Notas:
- VIES es un servicio externo y puede limitar automatizaciones.
- Este programa intenta ser rápido y amable: batches finitos + reintentos manuales.
"""

from __future__ import annotations

import re
import time
import random
import threading
import webbrowser
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, Optional, Tuple, List

import tkinter as tk
from tkinter import filedialog, messagebox

import ttkbootstrap as ttk

import requests
from openpyxl import load_workbook, Workbook
from zeep import Client
from zeep.exceptions import Fault, TransportError
from zeep.transports import Transport


CountryNumber = Tuple[str, str]  # (country, number)


class Tooltip:
    """Simple tooltip helper for ttk widgets."""
    def __init__(self, widget, text: str, delay_ms: int = 400):
        self.widget = widget
        self.text = text
        self.delay_ms = delay_ms
        self.tipwindow = None
        self.id = None
        
        widget.bind("<Enter>", self.on_enter, add=True)
        widget.bind("<Leave>", self.on_leave, add=True)
    
    def on_enter(self, event):
        self.schedule()
    
    def on_leave(self, event):
        self.unschedule()
        self.hidetip()
    
    def schedule(self):
        self.unschedule()
        self.id = self.widget.after(self.delay_ms, self.showtip)
    
    def unschedule(self):
        if self.id:
            self.widget.after_cancel(self.id)
            self.id = None
    
    def showtip(self):
        if self.tipwindow:
            return
        
        x = self.widget.winfo_rootx() + self.widget.winfo_width() // 2
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 5
        
        self.tipwindow = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        tw.attributes("-topmost", True)
        
        label = tk.Label(tw, text=self.text, background="#fff9e6", foreground="#333",
                         relief="solid", borderwidth=1, font=("Segoe UI", 8),
                         padx=8, pady=4, wraplength=250)
        label.pack()
    
    def hidetip(self):
        if self.tipwindow:
            self.tipwindow.destroy()
            self.tipwindow = None


@dataclass
class VatInfo:
    vat_clean: str
    country: str
    number: str
    nombre_excel: str = ""

    status: str = "NEW"  # NEW | VALIDATING | VALID | INVALID | THROTTLED | TIMEOUT | ERROR | PENDING_MAX | INVALID_FORMAT
    vies_name: str = ""
    vies_address: str = ""

    attempts_hard: int = 0
    throttles: int = 0

    last_checked_at: str = ""
    last_error: str = ""

    next_retry_at: Optional[datetime] = None

    # Anti-bucle / UX rápida: límites de reintento automático (por VAT)
    first_attempt_at: Optional[datetime] = None
    auto_retry_count: int = 0


class VATValidatorApp:
    VIES_WSDL = "https://ec.europa.eu/taxation_customs/vies/checkVatService.wsdl"
    VIES_WEB = "https://ec.europa.eu/taxation_customs/vies/"  # página pública

    TIMEOUT = 10
    MAX_ATTEMPTS = 3
    MAX_WORKERS = 3

    # Reintento automático *corto* (no bloquea la UI ni se queda en bucle)
    AUTO_RETRY_ENABLED = True
    AUTO_RETRY_MAX = 2              # reintentos automáticos por VAT (además del primer intento)
    AUTO_RETRY_DEADLINE_SEC = 25    # si en 25s no sale, se marca para manual

    THROTTLE_MS = 250  # separación mínima entre requests

    # Estados
    PENDING_STATES = {"NEW", "VALIDATING", "THROTTLED", "TIMEOUT", "ERROR", "PENDING_MAX"}
    VALIDATED_STATES = {"VALID", "INVALID"}

    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("VIES VAT Validator")
        self.root.geometry("1250x760")
        self.root.minsize(1100, 650)

        self.selected_file: Optional[Path] = None
        self.processing = False
        self._worker_threads: List[threading.Thread] = []
        self._stop_event = threading.Event()

        # Modelo
        self.vat_data: Dict[CountryNumber, VatInfo] = {}
        self._cache: Dict[CountryNumber, VatInfo] = {}  # cache en memoria por sesión (solo VALID/INVALID)

        # Concurrencia / throttle
        self._active_countries: set[str] = set()
        self._active_countries_lock = threading.Lock()
        self._throttle_lock = threading.Lock()
        self._last_request_time = 0.0

        # Cooldown por país (mini circuit-breaker)
        self._country_cooldown_until: Dict[str, float] = {}

        # Reutilización de conexiones por hilo (reduce TIMEOUTs)
        self._thread_local = threading.local()

        # UI state
        self.status_var = tk.StringVar(value="Listo")
        self.log_autoscroll_var = tk.BooleanVar(value=True)

        # Search vars
        self.search_pending_var = tk.StringVar(value="")
        self.search_validated_var = tk.StringVar(value="")

        # Map iids per tree
        self.pending_tree_iids: Dict[CountryNumber, str] = {}
        self.validated_tree_iids: Dict[CountryNumber, str] = {}

        # Stack for undo last validated
        self.undo_stack: List[Tuple[CountryNumber, str]] = []  # (key, previous_status)

        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

        self.apply_theme()
        self.setup_ui()

        self.log_info("Aplicación iniciada. Carga un Excel para comenzar.")

    # -------------------------
    # THEME / UI
    # -------------------------

    def apply_theme(self) -> None:
        style = ttk.Style()
        
        # Slango colors: soft blue background + white cards
        self.BG_MAIN = "#E8F0FE"  # soft blue background
        self.HEADER_BG = "#0b3a78"  # navy blue header
        self.CARD_BG = "#FFFFFF"  # white cards
        self.CARD_BORDER = "#E6E8EF"  # subtle border
        
        self.root.configure(bg=self.BG_MAIN)
        
        # Typography
        style.configure("Treeview", rowheight=28, background=self.CARD_BG, fieldbackground=self.CARD_BG, foreground="#1f2937")
        style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"), background="#f0f2f5", foreground="#1f2937")
        style.configure("TNotebook", background=self.CARD_BG)
        style.configure("TNotebook.Tab", padding=(14, 8))
        
        # Zebra striping: very soft tones
        self.TREE_ROW_EVEN = "#fafbfc"
        self.TREE_ROW_ODD = "#ffffff"
        self.LOG_BG = "#fafbfc"
        self.LOG_FG = "#1f2937"

    def setup_ui(self) -> None:
        # Root grid
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        self.root.rowconfigure(1, weight=0)

        # Main container (soft blue background)
        main = tk.Frame(self.root, bg=self.BG_MAIN)
        main.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)
        main.columnconfigure(0, weight=1)
        main.rowconfigure(2, weight=1)  # content
        main.rowconfigure(3, weight=0)  # log

        # Header (navy blue background, white text)
        header = tk.Frame(main, bg=self.HEADER_BG)
        header.grid(row=0, column=0, sticky="ew", pady=(0, 12), padx=0)
        header.columnconfigure(0, weight=1)

        title = tk.Label(header, text="VIES VAT Validator", font=("Segoe UI", 22, "bold"), bg=self.HEADER_BG, fg="#ffffff")
        title.grid(row=0, column=0, sticky="w", padx=16, pady=(12, 0))
        subtitle = tk.Label(header, text="Validación masiva de números VAT europeos", font=("Segoe UI", 10), bg=self.HEADER_BG, fg="#c7d2fe")
        subtitle.grid(row=1, column=0, sticky="w", padx=16, pady=(2, 12))

        # Toolbar (simplified)
        toolbar = tk.Frame(main, bg=self.BG_MAIN)
        toolbar.grid(row=1, column=0, sticky="ew", pady=(12, 8), padx=16)
        toolbar.columnconfigure(4, weight=1)

        self.load_btn = ttk.Button(toolbar, text="Cargar Excel", bootstyle="primary", padding=(14, 8), command=self.load_excel)
        self.load_btn.grid(row=0, column=0, padx=(0, 8))

        self.validate_btn = ttk.Button(toolbar, text="Validar", bootstyle="primary", padding=(14, 8), command=self.start_validation)
        self.validate_btn.grid(row=0, column=1, padx=(0, 8))
        self.validate_btn.state(["disabled"])

        self.retry_btn = ttk.Button(toolbar, text="Reintentar", bootstyle="primary", padding=(14, 8), command=self.retry_pending)
        self.retry_btn.grid(row=0, column=2, padx=(0, 8))
        self.retry_btn.state(["disabled"])
        Tooltip(self.retry_btn, "Reintenta los pendientes que ya han cumplido el tiempo de espera.")

        self.validate_selected_btn = ttk.Button(toolbar, text="Validar seleccionado", bootstyle="primary", padding=(14, 8), command=self.validate_selected)
        self.validate_selected_btn.grid(row=0, column=3, padx=(0, 8))
        self.validate_selected_btn.state(["disabled"])
        Tooltip(self.validate_selected_btn, "Valida únicamente el VAT seleccionado.")

        self.undo_btn = ttk.Button(toolbar, text="Deshacer último", bootstyle="secondary-outline", padding=(10, 6), command=self.undo_last_validated)
        self.undo_btn.grid(row=0, column=4, padx=(0, 8))
        self.undo_btn.state(["disabled"])

        # Export menu button (right side)
        self.export_menu_btn = ttk.Menubutton(toolbar, text="Exportar ▾", bootstyle="secondary-outline")
        self.export_menu_btn.grid(row=0, column=5, sticky="e", padx=(0, 0))
        self.export_menu_btn.state(["disabled"])
        
        self.export_menu = tk.Menu(self.export_menu_btn, tearoff=0)
        self.export_menu.add_command(label="Exportar todo", command=lambda: self.export_to_excel(scope="all"))
        self.export_menu.add_command(label="Exportar validados", command=lambda: self.export_to_excel(scope="validated"))
        self.export_menu.add_command(label="Exportar pendientes", command=lambda: self.export_to_excel(scope="pending"))
        self.export_menu_btn["menu"] = self.export_menu

        # Content area
        content = tk.Frame(main, bg=self.BG_MAIN)
        content.grid(row=2, column=0, sticky="nsew")
        content.columnconfigure(0, weight=1)
        content.rowconfigure(0, weight=1)

        # Results card (white frame, no Labelframe)
        results_card = tk.Frame(content, bg=self.CARD_BG, relief="solid", borderwidth=1, highlightbackground=self.CARD_BORDER, highlightthickness=1)
        results_card.grid(row=0, column=0, sticky="nsew", pady=(0, 12), padx=16)
        results_card.columnconfigure(0, weight=1)
        results_card.rowconfigure(2, weight=1)
        
        # Title label on white background
        tk.Label(results_card, text="Resultados", font=("Segoe UI", 12, "bold"), bg=self.CARD_BG, fg="#1f2937", anchor="w", padx=12, pady=8).grid(row=0, column=0, sticky="ew")

        # Banner for pending VATs (will be shown/hidden dynamically)
        # Banner for pending VATs (Slango colors: cyan + blue)
        self.banner_frame = tk.Frame(results_card, bg="#E6F6F8", relief="solid", borderwidth=1, highlightbackground="#2FB5C4", highlightthickness=1)
        self.banner_frame.grid(row=1, column=0, sticky="ew", padx=8, pady=8)
        self.banner_frame.columnconfigure(0, weight=1)
        self.banner_frame.grid_remove()  # Hidden by default
        
        # Banner content
        banner_content = tk.Frame(self.banner_frame, bg="#E6F6F8")
        banner_content.grid(row=0, column=0, sticky="ew", padx=12, pady=8)
        banner_content.columnconfigure(0, weight=1)
        
        self.banner_label = tk.Label(banner_content, text="", font=("Segoe UI", 9), bg="#E6F6F8", fg="#0F3A6D")
        self.banner_label.grid(row=0, column=0, sticky="w")
        
        banner_btn_frame = tk.Frame(banner_content, bg="#E6F6F8")
        banner_btn_frame.grid(row=0, column=1, sticky="e", padx=(12, 0))
        
        # Custom styled buttons for banner
        self.banner_go_pending_btn = tk.Button(banner_btn_frame, text="Ir a Pendientes", bg="#2FB5C4", fg="white", font=("Segoe UI", 9, "bold"), relief="solid", borderwidth=0, padx=10, pady=4, cursor="hand2", command=self._go_to_pending_tab, activebackground="#239EAC", activeforeground="white")
        self.banner_go_pending_btn.grid(row=0, column=0, padx=(0, 6))
        
        self.banner_retry_btn = tk.Button(banner_btn_frame, text="Reintentar ahora", bg="#2FB5C4", fg="white", font=("Segoe UI", 9, "bold"), relief="solid", borderwidth=0, padx=10, pady=4, cursor="hand2", command=self.retry_pending, activebackground="#239EAC", activeforeground="white")
        self.banner_retry_btn.grid(row=0, column=1)
        Tooltip(self.banner_retry_btn, "Reintenta los VAT que estén listos en este momento.")

        self.notebook = ttk.Notebook(results_card)
        self.notebook.grid(row=2, column=0, sticky="nsew", padx=12, pady=(0, 12))

        # Tabs
        self.pending_tab = ttk.Frame(self.notebook, padding=(10, 10))
        self.validated_tab = ttk.Frame(self.notebook, padding=(10, 10))
        self.notebook.add(self.pending_tab, text="Pendientes")
        self.notebook.add(self.validated_tab, text="Validados")

        self._build_tab(self.pending_tab, kind="pending")
        self._build_tab(self.validated_tab, kind="validated")

        # Log card (white frame, no Labelframe)
        log_card = tk.Frame(main, bg=self.CARD_BG, relief="solid", borderwidth=1, highlightbackground=self.CARD_BORDER, highlightthickness=1)
        log_card.grid(row=3, column=0, sticky="ew", padx=16, pady=(0, 12))
        log_card.columnconfigure(0, weight=1)
        log_card.rowconfigure(1, weight=1)

        # Header with title and controls
        header_row = tk.Frame(log_card, bg=self.CARD_BG)
        header_row.grid(row=0, column=0, sticky="ew", padx=12, pady=8)
        header_row.columnconfigure(0, weight=1)
        
        tk.Label(header_row, text="Registro de actividad", font=("Segoe UI", 12, "bold"), bg=self.CARD_BG, fg="#1f2937", anchor="w").grid(row=0, column=0, sticky="w")

        # Controls right
        controls = tk.Frame(header_row, bg=self.CARD_BG)
        controls.grid(row=0, column=1, sticky="e")

        self.autoscroll_chk = ttk.Checkbutton(controls, text="Auto-scroll", variable=self.log_autoscroll_var, bootstyle="round-toggle")
        self.autoscroll_chk.grid(row=0, column=0, padx=(0, 10))

        ttk.Button(controls, text="Limpiar", bootstyle="secondary-outline", command=self.clear_log).grid(row=0, column=1, padx=(0, 8))
        ttk.Button(controls, text="Copiar", bootstyle="secondary-outline", command=self.copy_log).grid(row=0, column=2)

        # Text + scrollbar (reduced height)
        text_frame = tk.Frame(log_card, bg=self.CARD_BG)
        text_frame.grid(row=1, column=0, sticky="ew", padx=12, pady=(0, 12))
        text_frame.columnconfigure(0, weight=1)

        self.log_text = tk.Text(
            text_frame,
            height=5,
            wrap=tk.WORD,
            font=("Consolas", 9),
            bg=self.LOG_BG,
            fg=self.LOG_FG,
            insertbackground=self.LOG_FG,
            relief="flat",
            highlightthickness=1,
            highlightbackground="#e6e8ef",
            highlightcolor="#b6c2cf",
        )
        self.log_text.grid(row=0, column=0, sticky="ew")

        log_scroll = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=self.log_text.yview, bootstyle="round")
        log_scroll.grid(row=0, column=1, sticky="ns")
        self.log_text.configure(yscrollcommand=log_scroll.set)
        self.log_text.configure(state=tk.DISABLED)

        # Tags
        self.log_text.tag_configure("OK", foreground="#10b981")
        self.log_text.tag_configure("WARN", foreground="#f59e0b")
        self.log_text.tag_configure("ERROR", foreground="#ef4444")
        self.log_text.tag_configure("INFO", foreground="#1f2937")
        self.log_text.tag_configure("DEBUG", foreground="#6b7280")

        # Auto-scroll inteligente: si el usuario sube, lo apagamos
        self._install_log_scroll_detection(log_scroll)

        # Footer (status + exit)
        footer = tk.Frame(self.root, bg=self.BG_MAIN)
        footer.grid(row=1, column=0, sticky="ew", padx=20, pady=10)
        footer.columnconfigure(0, weight=1)

        self.status_label = tk.Label(footer, textvariable=self.status_var, anchor="w", bg=self.BG_MAIN, fg="#1f2937", font=("Segoe UI", 9), padx=8)
        self.status_label.grid(row=0, column=0, sticky="ew")

        self.exit_btn = ttk.Button(footer, text="Salir", bootstyle="danger-outline", width=12, command=self.exit_app)
        self.exit_btn.grid(row=0, column=1, sticky="e", padx=(10, 0), pady=(6, 4))

    def _build_tab(self, parent: ttk.Frame, kind: str) -> None:
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(1, weight=1)

        # Search
        row = ttk.Frame(parent)
        row.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        ttk.Label(row, text="Buscar:").grid(row=0, column=0, sticky="w")

        if kind == "pending":
            entry = ttk.Entry(row, textvariable=self.search_pending_var)
            entry.grid(row=0, column=1, sticky="ew", padx=(8, 0))
            row.columnconfigure(1, weight=1)
            entry.bind("<KeyRelease>", lambda _e: self.refresh_trees())
        else:
            entry = ttk.Entry(row, textvariable=self.search_validated_var)
            entry.grid(row=0, column=1, sticky="ew", padx=(8, 0))
            row.columnconfigure(1, weight=1)
            entry.bind("<KeyRelease>", lambda _e: self.refresh_trees())

        # Tree
        tree_frame = ttk.Frame(parent)
        tree_frame.grid(row=1, column=0, sticky="nsew")
        tree_frame.columnconfigure(0, weight=1)
        tree_frame.rowconfigure(0, weight=1)

        # Define columns based on kind
        if kind == "pending":
            cols = ("VAT", "País", "Número", "Nombre", "Estado", "Intentos", "Última verificación", "Siguiente intento", "Error", "Acción")
        else:
            cols = ("VAT", "País", "Número", "Nombre", "Estado", "Última verificación")

        tree = ttk.Treeview(tree_frame, columns=cols, show="headings", height=12)
        tree.grid(row=0, column=0, sticky="nsew")

        # Configure headings and columns
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=120, anchor="w")

        # Adjust column widths
        tree.column("VAT", width=140)
        tree.column("País", width=60, anchor="center")
        tree.column("Número", width=120)
        tree.column("Nombre", width=220)
        tree.column("Estado", width=140)
        
        if kind == "pending":
            tree.column("Intentos", width=80, anchor="center")
            tree.column("Última verificación", width=150)
            tree.column("Siguiente intento", width=130)
            tree.column("Error", width=210)
            tree.column("Acción", width=120, anchor="center")
        else:
            tree.column("Última verificación", width=150)

        yscroll = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview, bootstyle="round")
        yscroll.grid(row=0, column=1, sticky="ns")
        tree.configure(yscrollcommand=yscroll.set)

        xscroll = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=tree.xview, bootstyle="round")
        xscroll.grid(row=1, column=0, sticky="ew")
        tree.configure(xscrollcommand=xscroll.set)

        # Zebra striping
        tree.tag_configure("row_even", background=self.TREE_ROW_EVEN)
        tree.tag_configure("row_odd", background=self.TREE_ROW_ODD)

        # Row tags (colors)
        tree.tag_configure("VALID", foreground="#0f766e")
        tree.tag_configure("INVALID", foreground="#b91c1c")
        tree.tag_configure("PENDING", foreground="#b45309")
        tree.tag_configure("THROTTLED", foreground="#b45309")
        tree.tag_configure("TIMEOUT", foreground="#b45309")
        tree.tag_configure("ERROR", foreground="#b45309")
        tree.tag_configure("PENDING_MAX", foreground="#64748b")
        tree.tag_configure("INVALID_FORMAT", foreground="#64748b")

        # Bindings
        tree.bind("<<TreeviewSelect>>", self.on_tree_select)
        tree.bind("<Button-1>", self.on_tree_click)
        tree.bind("<Double-1>", self.on_tree_double_click)
        tree.bind("<Motion>", self.on_tree_motion)
        tree.bind("<Button-3>", self.show_context_menu)

        if kind == "pending":
            self.pending_tree = tree
        else:
            self.validated_tree = tree

        # Context menu (minimal)
        # Nota: se construye una sola vez y se reutiliza.
        self.tree_context_menu = tk.Menu(self.root, tearoff=0)
        self.tree_context_menu.add_command(label="Abrir en VIES (web)", command=self.open_vies_web)
        self.tree_context_menu.add_separator()
        self.tree_context_menu.add_command(label="Copiar número VAT", command=lambda: self.copy_vat(number_only=True))
        self.tree_context_menu.add_command(label="Copiar VAT completo", command=lambda: self.copy_vat(number_only=False))

    def _install_log_scroll_detection(self, scrollbar: ttk.Scrollbar) -> None:
        # Detectar cuando el usuario se separa del final: desactivar auto-scroll.
        def on_scroll(*args):
            self.log_text.yview(*args)
            self._check_log_autoscroll_state()

        scrollbar.configure(command=on_scroll)

        def on_mousewheel(event):
            # Windows uses delta, other platforms may differ.
            try:
                self.log_text.yview_scroll(int(-1 * (event.delta / 120)), "units")
            except Exception:
                return
            self._check_log_autoscroll_state()
            return "break"

        self.log_text.bind("<MouseWheel>", on_mousewheel)

    def _check_log_autoscroll_state(self) -> None:
        # yview returns (first, last) in [0,1]
        _, last = self.log_text.yview()
        # If not at bottom, auto-scroll OFF
        if last < 0.999:
            self.log_autoscroll_var.set(False)

    # -------------------------
    # STATUS / LOG
    # -------------------------

    def set_status(self, msg: str, timeout_ms: int = 4000) -> None:
        self.status_var.set(msg)
        self.status_label.update_idletasks()
        if timeout_ms > 0:
            self.root.after(timeout_ms, lambda: self.status_var.set("Listo"))

    def _log(self, level: str, message: str) -> None:
        ts = datetime.now().strftime("%H:%M:%S")
        line = f"[{ts}] {message}\n"
        self.log_text.configure(state=tk.NORMAL)
        self.log_text.insert(tk.END, line, level)
        if self.log_autoscroll_var.get():
            self.log_text.see(tk.END)
        self.log_text.configure(state=tk.DISABLED)

    def log_ok(self, msg: str) -> None:
        self._log("OK", msg)

    def log_warn(self, msg: str) -> None:
        self._log("WARN", msg)

    def log_error(self, msg: str) -> None:
        self._log("ERROR", msg)

    def log_info(self, msg: str) -> None:
        self._log("INFO", msg)

    def log_debug(self, msg: str) -> None:
        self._log("DEBUG", msg)

    def clear_log(self) -> None:
        self.log_text.configure(state=tk.NORMAL)
        self.log_text.delete("1.0", tk.END)
        self.log_text.configure(state=tk.DISABLED)
        self.set_status("Registro limpiado", 2000)

    def copy_log(self) -> None:
        content = self.log_text.get("1.0", tk.END)
        self.root.clipboard_clear()
        self.root.clipboard_append(content)
        self.set_status("Registro copiado al portapapeles", 2500)

    # -------------------------
    # VAT helpers
    # -------------------------

    @staticmethod
    def normalize_vat(vat) -> Optional[str]:
        if vat is None:
            return None
        vat_str = str(vat).replace("\u00A0", "")
        vat_clean = re.sub(r"[^A-Z0-9]", "", vat_str.upper())
        return vat_clean or None

    def parse_vat(self, vat) -> Tuple[Optional[str], Optional[str], Optional[str]]:
        normalized = self.normalize_vat(vat)
        if not normalized or len(normalized) < 3:
            return None, None, normalized
        if not re.match(r"^[A-Z]{2}", normalized):
            return None, None, normalized
        return normalized[:2], normalized[2:], normalized

    @staticmethod
    def get_vat_number_only(vat_clean: str) -> str:
        vat_clean = re.sub(r"[^A-Z0-9]", "", (vat_clean or "").upper())
        if re.match(r"^[A-Z]{2}", vat_clean):
            return vat_clean[2:]
        return vat_clean

    def human_status(self, status: str) -> str:
        mapping = {
            "VALID": "✓ Válido",
            "INVALID": "✕ Inválido",
            "NEW": "⏳ Pendiente",
            "VALIDATING": "⏳ Pendiente",
            "THROTTLED": "⛔ Limitado por VIES",
            "TIMEOUT": "… Sin respuesta",
            "ERROR": "⚠ Error",
            "PENDING_MAX": "⚠ No verificable ahora",
            "INVALID_FORMAT": "✕ Formato inválido",
        }
        return mapping.get(status, status)

    def accion_text(self, status: str) -> str:
        if status in {"THROTTLED", "TIMEOUT", "ERROR", "PENDING_MAX"}:
            return "[[ Abrir VIES ]]"
        return ""

    # -------------------------
    # UI interactions
    # -------------------------

    def _active_tree(self) -> ttk.Treeview:
        # active tab tree: pending or validated
        tab = self.notebook.index(self.notebook.select())
        return self.pending_tree if tab == 0 else self.validated_tree

    def on_tree_select(self, _event=None) -> None:
        tree = self._active_tree()
        sel = tree.selection()
        if sel and not self.processing:
            self.validate_selected_btn.state(["!disabled"])
        else:
            self.validate_selected_btn.state(["disabled"])

    def on_tree_motion(self, event) -> None:
        tree = event.widget
        region = tree.identify_region(event.x, event.y)
        if region != "cell":
            tree.configure(cursor="")
            return
        col = tree.identify_column(event.x)
        row = tree.identify_row(event.y)
        if not row:
            tree.configure(cursor="")
            return
        # Acción is last column
        if col == f"#{len(tree['columns'])}":
            vals = tree.item(row, "values")
            if vals and str(vals[-1]).strip():
                tree.configure(cursor="hand2")
                return
        tree.configure(cursor="")

    def on_tree_click(self, event) -> None:
        tree = event.widget
        region = tree.identify_region(event.x, event.y)
        if region != "cell":
            return
        col = tree.identify_column(event.x)
        row = tree.identify_row(event.y)
        if not row:
            return
        # Acción column
        if col == f"#{len(tree['columns'])}":
            vals = tree.item(row, "values")
            if vals and str(vals[-1]).strip():
                tree.selection_set(row)
                self.open_vies_web()

    def on_tree_double_click(self, event) -> None:
        tree = event.widget
        row = tree.identify_row(event.y)
        if not row:
            return
        vals = tree.item(row, "values")
        if vals and str(vals[-1]).strip():
            tree.selection_set(row)
            self.open_vies_web()

    def show_context_menu(self, event) -> None:
        tree = event.widget
        row = tree.identify_row(event.y)
        if not row:
            return
        tree.selection_set(row)
        self.tree_context_menu.tk_popup(event.x_root, event.y_root)

    def copy_vat(self, number_only: bool) -> None:
        sel = self._get_selected_key()
        if not sel:
            return
        info = self.vat_data.get(sel)
        if not info:
            return
        text = self.get_vat_number_only(info.vat_clean) if number_only else info.vat_clean
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        if number_only:
            self.set_status(f"Copiado: {text} (País: {info.country})", 3500)
        else:
            self.set_status(f"Copiado: {text}", 2500)

    def open_vies_web(self) -> None:
        sel = self._get_selected_key()
        if not sel:
            return
        info = self.vat_data.get(sel)
        if not info:
            return
        number_only = self.get_vat_number_only(info.vat_clean)

        def do_ui():
            self.root.clipboard_clear()
            self.root.clipboard_append(number_only)
            
            # Show confirmation dialog
            dialog_msg = (
                f"Se ha copiado al portapapeles el número: {number_only}\n"
                f"País: {info.country}\n\n"
                f"En VIES selecciona el país {info.country} y pega el número.\n\n"
                f"¿Quieres abrir VIES ahora?"
            )
            
            status_msg = f"Copiado: {number_only} (País {info.country}). Confirmación para abrir VIES."
            self.set_status(status_msg, 3500)
            
            # Ask user for confirmation
            if messagebox.askokcancel("Abrir VIES", dialog_msg):
                webbrowser.open(self.VIES_WEB)

        # Execute in UI thread
        self.root.after(0, do_ui)

    # -------------------------
    # Loading / rendering
    # -------------------------

    def refresh_trees(self) -> None:
        # Re-render both trees based on filters
        self._render_pending_tree()
        self._render_validated_tree()

    def _render_pending_tree(self) -> None:
        tree = self.pending_tree
        tree.delete(*tree.get_children())
        self.pending_tree_iids.clear()

        q = self.search_pending_var.get().strip().lower()

        for idx, (key, info) in enumerate(self.vat_data.items()):
            if info.status in self.VALIDATED_STATES:
                continue
            if info.status == "INVALID_FORMAT":
                # keep in pending tab
                pass

            if q:
                hay = f"{info.vat_clean} {info.nombre_excel}".lower()
                if q not in hay:
                    continue

            attempts = info.attempts_hard + info.throttles
            last_checked = info.last_checked_at
            retry = ""
            if info.next_retry_at:
                retry = info.next_retry_at.strftime("%H:%M:%S")

            values = (
                info.vat_clean,
                info.country,
                info.number,
                info.nombre_excel,
                self.human_status(info.status),
                attempts,
                last_checked,
                retry,
                info.last_error,
                self.accion_text(info.status),
            )
            iid = str(key)
            row_tag = "row_even" if idx % 2 == 0 else "row_odd"
            tree.insert("", "end", iid=iid, values=values, tags=(row_tag, info.status))
            self.pending_tree_iids[key] = iid

    def _render_validated_tree(self) -> None:
        tree = self.validated_tree
        tree.delete(*tree.get_children())
        self.validated_tree_iids.clear()

        q = self.search_validated_var.get().strip().lower()

        for idx, (key, info) in enumerate(self.vat_data.items()):
            if info.status not in self.VALIDATED_STATES:
                continue
            if q:
                hay = f"{info.vat_clean} {info.nombre_excel} {info.vies_name}".lower()
                if q not in hay:
                    continue

            last_checked = info.last_checked_at
            # Only 6 columns for validated tree
            values = (
                info.vat_clean,
                info.country,
                info.number,
                info.nombre_excel or info.vies_name,
                self.human_status(info.status),
                last_checked,
            )
            iid = str(key)
            row_tag = "row_even" if idx % 2 == 0 else "row_odd"
            tree.insert("", "end", iid=iid, values=values, tags=(row_tag, info.status))
            self.validated_tree_iids[key] = iid

    def _get_selected_key(self) -> Optional[CountryNumber]:
        tree = self._active_tree()
        sel = tree.selection()
        if not sel:
            return None
        iid = sel[0]
        try:
            # iid is str(key) where key is tuple; recover by searching
            vals = tree.item(iid, "values")
            if not vals:
                return None
            country = vals[1]
            number = vals[2]
            return (country, number)
        except Exception:
            return None

    # -------------------------
    # Excel load / export
    # -------------------------

    def load_excel(self) -> None:
        if self.processing:
            messagebox.showinfo("Info", "Hay una validación en curso.")
            return

        file_path = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
        )
        if not file_path:
            return

        self.selected_file = Path(file_path)
        self.vat_data.clear()

        # UI state
        self.retry_btn.state(["disabled"])
        self.export_menu_btn.state(["disabled"])
        self.validate_selected_btn.state(["disabled"])

        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active

            # Detect header row: find columns containing NIF/VAT
            header_row = None
            nif_col = None
            name_col = None

            for r in range(1, min(30, ws.max_row) + 1):
                row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
                row_text = " ".join([str(v).strip().upper() for v in row_vals if v is not None])
                if any(k in row_text for k in ["NIF", "VAT", "VAT NUMBER", "VAT_NUMBER"]):
                    header_row = r
                    # Find column indices
                    for c in range(1, ws.max_column + 1):
                        val = ws.cell(row=r, column=c).value
                        if val is None:
                            continue
                        t = str(val).strip().upper()
                        if t in {"NIF", "VAT", "VAT NUMBER", "VAT_NUMBER"}:
                            nif_col = c
                        if t in {"NOMBRE", "NAME", "RAZON SOCIAL", "RAZÓN SOCIAL"}:
                            name_col = c
                    break

            if header_row is None or nif_col is None:
                # fallback: if first column looks like VAT values
                header_row = 1
                nif_col = 1

            # Load data
            for r in range(header_row + 1, ws.max_row + 1):
                raw_vat = ws.cell(row=r, column=nif_col).value
                if raw_vat is None:
                    continue

                country, number, vat_clean = self.parse_vat(raw_vat)
                nombre_excel = ""
                if name_col:
                    v = ws.cell(row=r, column=name_col).value
                    if v is not None:
                        nombre_excel = str(v).strip()

                if not vat_clean:
                    continue

                if not country or not number:
                    # invalid format
                    # Use placeholder country/number to keep stable key: try to split anyway
                    # We'll store under ("", vat_clean)
                    key = ("", vat_clean)
                    self.vat_data[key] = VatInfo(vat_clean=vat_clean, country="", number="", nombre_excel=nombre_excel, status="INVALID_FORMAT", last_error="Formato inválido")
                    continue

                key = (country, number)

                # Cache hit (VALID/INVALID): reuse
                cached = self._cache.get(key)
                if cached and cached.status in self.VALIDATED_STATES:
                    info = VatInfo(
                        vat_clean=vat_clean,
                        country=country,
                        number=number,
                        nombre_excel=nombre_excel or cached.nombre_excel,
                        status=cached.status,
                        vies_name=cached.vies_name,
                        vies_address=cached.vies_address,
                        attempts_hard=cached.attempts_hard,
                        throttles=cached.throttles,
                        last_checked_at=cached.last_checked_at,
                        last_error=cached.last_error,
                    )
                    self.vat_data[key] = info
                    continue

                self.vat_data[key] = VatInfo(vat_clean=vat_clean, country=country, number=number, nombre_excel=nombre_excel)

            self.refresh_trees()

            self.log_ok(f"Cargados {len(self.vat_data)} VATs únicos desde {self.selected_file.name}")
            self.set_status(f"Excel cargado: {len(self.vat_data)} VATs. Pulsa 'Validar' para comenzar.", 3500)

            # Enable controls
            self.validate_btn.state(["!disabled"])
            self.export_menu_btn.state(["!disabled"])

        except Exception as e:
            self.log_error(f"Error al cargar Excel: {e}")
            messagebox.showerror("Error", f"No se pudo cargar el archivo: {e}")

    def export_to_excel(self, scope: str = "all") -> None:
        if not self.vat_data:
            messagebox.showinfo("Info", "No hay datos para exportar")
            return

        if scope not in {"all", "pending", "validated"}:
            scope = "all"

        initial_name = "vat_results.xlsx"
        if self.selected_file:
            suffix = {"all": "validated", "pending": "pending", "validated": "validated"}[scope]
            initial_name = f"{self.selected_file.stem}_{suffix}.xlsx"

        out = filedialog.asksaveasfilename(
            title="Guardar resultados",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("All", "*.*")],
            initialfile=initial_name,
        )
        if not out:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "VAT Results"

            headers = [
                "VAT_CLEAN",
                "COUNTRY",
                "NUMBER",
                "NOMBRE_EXCEL",
                "STATUS",
                "ATTEMPTS",
                "THROTTLES",
                "LAST_CHECKED_AT",
                "NEXT_RETRY_AT",
                "VIES_NAME",
                "VIES_ADDRESS",
                "ERROR",
            ]
            for i, h in enumerate(headers, start=1):
                ws.cell(row=1, column=i, value=h)

            row = 2
            for key, info in self.vat_data.items():
                if scope == "pending" and info.status in self.VALIDATED_STATES:
                    continue
                if scope == "validated" and info.status not in self.VALIDATED_STATES:
                    continue

                attempts = info.attempts_hard + info.throttles
                ws.cell(row=row, column=1, value=info.vat_clean)
                ws.cell(row=row, column=2, value=info.country)
                ws.cell(row=row, column=3, value=info.number)
                ws.cell(row=row, column=4, value=info.nombre_excel)
                ws.cell(row=row, column=5, value=info.status)
                ws.cell(row=row, column=6, value=attempts)
                ws.cell(row=row, column=7, value=info.throttles)
                ws.cell(row=row, column=8, value=info.last_checked_at)
                ws.cell(row=row, column=9, value=info.next_retry_at.strftime("%Y-%m-%d %H:%M:%S") if info.next_retry_at else "")
                ws.cell(row=row, column=10, value=info.vies_name)
                ws.cell(row=row, column=11, value=info.vies_address)
                ws.cell(row=row, column=12, value=info.last_error)
                row += 1

            wb.save(out)
            self.log_ok(f"Exportado: {Path(out).name}")
            self.set_status(f"Exportado: {Path(out).name}", 3500)
            messagebox.showinfo("Éxito", f"Resultados exportados a:\n{out}")
        except Exception as e:
            self.log_error(f"Error al exportar: {e}")
            messagebox.showerror("Error", f"No se pudo exportar: {e}")

    # -------------------------
    # Validation flow
    # -------------------------

    def start_validation(self) -> None:
        if self.processing:
            return

        to_validate = [(k, v) for k, v in self.vat_data.items() if v.status == "NEW" and v.country and v.number]
        if not to_validate:
            return

        self.processing = True
        self._stop_event.clear()
        self.load_btn.state(["disabled"])
        self.validate_btn.state(["disabled"])
        self.retry_btn.state(["disabled"])
        self.validate_selected_btn.state(["disabled"])

        self.log_info("============================================================")
        self.log_info(f"Iniciando validación de {len(to_validate)} VATs...")

        t = threading.Thread(target=self._validate_batch_worker, args=(to_validate,), daemon=True)
        self._worker_threads = [t]
        t.start()

    def retry_pending(self) -> None:
        if self.processing:
            return

        now = datetime.now()
        to_retry = []
        next_retry_time = None

        for k, v in self.vat_data.items():
            if v.status in {"THROTTLED", "TIMEOUT", "ERROR", "PENDING_MAX"}:
                if v.next_retry_at is None or v.next_retry_at <= now:
                    to_retry.append((k, v))
                elif next_retry_time is None or v.next_retry_at < next_retry_time:
                    next_retry_time = v.next_retry_at

        if not to_retry:
            if next_retry_time:
                wait_secs = int((next_retry_time - now).total_seconds())
                msg = f"Aún no se puede reintentar. Próximo reintento recomendado en {wait_secs}s."
                messagebox.showinfo("Reintentar", msg)
            else:
                messagebox.showinfo("Info", "No hay VATs pendientes listos para reintentar.")
            return

        self.processing = True
        self._stop_event.clear()
        self.load_btn.state(["disabled"])
        self.retry_btn.state(["disabled"])
        self.validate_selected_btn.state(["disabled"])

        self.log_info("============================================================")
        self.log_info(f"Reintentando {len(to_retry)} VATs pendientes...")

        t = threading.Thread(target=self._validate_batch_worker, args=(to_retry,), daemon=True)
        self._worker_threads = [t]
        t.start()

    def validate_selected(self) -> None:
        if self.processing:
            return
        key = self._get_selected_key()
        if not key:
            messagebox.showinfo("Info", "Selecciona un VAT.")
            return

        info = self.vat_data.get(key)
        if not info or not info.country or not info.number:
            return

        # 1 intento inmediato
        self.processing = True
        self._stop_event.clear()
        self.load_btn.state(["disabled"])
        self.validate_btn.state(["disabled"])
        self.retry_btn.state(["disabled"])
        self.validate_selected_btn.state(["disabled"])

        self.log_info(f"Validando seleccionado: {info.vat_clean}")

        t = threading.Thread(target=self._validate_batch_worker, args=([(key, info)],), daemon=True)
        self._worker_threads = [t]
        t.start()

    def _validate_batch_worker(self, items: List[Tuple[CountryNumber, VatInfo]]) -> None:
        total = len(items)
        completed = 0

        # Keys that reached a terminal state inside this run
        finished: set[CountryNumber] = set()

        # Queue scheduler: (ready_time, counter, key)
        pending: List[Tuple[float, int, CountryNumber]] = []
        seq = 0
        seq_lock = threading.Lock()

        def next_seq() -> int:
            nonlocal seq
            with seq_lock:
                seq += 1
                return seq
        now = time.time()
        for k, _info in items:
            pending.append((now, next_seq(), k))

        pending.sort(key=lambda x: (x[0], x[1]))

        def pop_ready() -> Optional[CountryNumber]:
            nonlocal pending
            if not pending:
                return None
            # find first ready + available country
            current = time.time()
            scan = min(len(pending), 60)  # evita que 15 primeros bloqueen a otros países
            for idx, (ready, _c, key) in enumerate(pending[:scan]):
                info = self.vat_data.get(key)
                if not info:
                    continue
                if ready > current:
                    continue

                # cooldown por país (mini circuit breaker)
                cd_until = self._country_cooldown_until.get(info.country)
                if cd_until and cd_until > current:
                    continue

                with self._active_countries_lock:
                    if info.country in self._active_countries:
                        continue
                    self._active_countries.add(info.country)
                # remove
                pending.pop(idx)
                return key
            return None

        def mark_country_done(country: str) -> None:
            with self._active_countries_lock:
                self._active_countries.discard(country)

        def worker_loop():
            nonlocal completed
            while not self._stop_event.is_set():
                key = pop_ready()
                if key is None:
                    if not pending:
                        break
                    time.sleep(0.08)
                    continue

                info = self.vat_data.get(key)
                if not info:
                    continue

                try:
                    # throttle global
                    with self._throttle_lock:
                        elapsed = time.time() - self._last_request_time
                        min_gap = self.THROTTLE_MS / 1000.0
                        if elapsed < min_gap:
                            time.sleep(min_gap - elapsed)
                        self._last_request_time = time.time()

                    info.status = "VALIDATING"
                    info.last_error = ""
                    info.last_checked_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                    if info.first_attempt_at is None:
                        info.first_attempt_at = datetime.now()

                    self.root.after(0, self.refresh_trees)

                    result = self.validate_vat_with_vies(info.country, info.number)
                    retry_ready_at = self._apply_result(info, result)

                    # Si fue THROTTLED, ponemos un cooldown mínimo por país para no martillear
                    if info.status == "THROTTLED" and info.next_retry_at is not None:
                        self._country_cooldown_until[info.country] = max(
                            self._country_cooldown_until.get(info.country, 0.0),
                            info.next_retry_at.timestamp(),
                        )

                    # Auto requeue (corto + finito) para errores temporales
                    if (
                        self.AUTO_RETRY_ENABLED
                        and retry_ready_at is not None
                        and (info.country, info.number) not in finished
                    ):
                        # Deadline por VAT: si tarda demasiado, lo mandamos a manual
                        if info.first_attempt_at is not None:
                            elapsed = (datetime.now() - info.first_attempt_at).total_seconds()
                        else:
                            elapsed = 0

                        if elapsed <= self.AUTO_RETRY_DEADLINE_SEC and info.auto_retry_count < self.AUTO_RETRY_MAX:
                            info.auto_retry_count += 1
                            pending.append((retry_ready_at, next_seq(), (info.country, info.number)))
                            pending.sort(key=lambda x: (x[0], x[1]))
                            # No contamos como completado aún
                            continue
                        else:
                            # Se acabó el tiempo o el presupuesto: queda para manual
                            info.status = "PENDING_MAX"
                            info.next_retry_at = None
                            info.last_error = info.last_error or "Manual recomendado (VIES inestable)"
                            self.root.after(0, self.refresh_trees)

                    # Terminal states for this run
                    if info.status in self.VALIDATED_STATES or info.status in {"PENDING_MAX", "INVALID_FORMAT"}:
                        finished.add((info.country, info.number))

                    # Cache VALID/INVALID
                    if info.status in self.VALIDATED_STATES:
                        self._cache[(info.country, info.number)] = info

                    # Completed count
                    completed += 1
                    self.root.after(0, lambda c=completed: self.set_status(f"Validando… {c}/{total}", 0))

                finally:
                    mark_country_done(info.country)

        # Start workers
        threads = []
        for _ in range(min(self.MAX_WORKERS, max(1, total))):
            t = threading.Thread(target=worker_loop, daemon=True)
            t.start()
            threads.append(t)

        for t in threads:
            t.join()

        self.root.after(0, self._finish_validation)

    def _apply_result(self, info: VatInfo, result: dict) -> Optional[float]:
        status = result.get("status")
        now = datetime.now()
        info.last_checked_at = now.strftime("%Y-%m-%d %H:%M:%S")
        prev_status = info.status  # Save previous status for undo

        if status == "VALID":
            info.status = "VALID"
            info.vies_name = result.get("vies_name", "")
            info.vies_address = result.get("vies_address", "")
            info.last_error = ""
            info.next_retry_at = None
            self.log_ok(f"{info.vat_clean} → VALID")
            # Save to undo stack if transitioning to validated
            if prev_status not in self.VALIDATED_STATES:
                self.undo_stack.append(((info.country, info.number), prev_status))
                self.undo_btn.state(["!disabled"])

            self.root.after(0, self.refresh_trees)
            return None

        elif status == "INVALID":
            info.status = "INVALID"
            info.vies_name = ""
            info.vies_address = ""
            info.last_error = ""
            info.next_retry_at = None
            self.log_warn(f"{info.vat_clean} → INVALID")
            # Save to undo stack if transitioning to validated
            if prev_status not in self.VALIDATED_STATES:
                self.undo_stack.append(((info.country, info.number), prev_status))
                self.undo_btn.state(["!disabled"])

            self.root.after(0, self.refresh_trees)
            return None

        elif status == "THROTTLED":
            info.throttles += 1
            info.status = "THROTTLED"
            info.last_error = result.get("error", "MS_MAX_CONCURRENT_REQ")

            # Retry suggestion (no auto retry en el mismo batch)
            throttles = info.throttles
            if throttles == 1:
                jitter = random.uniform(2, 7)
            elif throttles == 2:
                jitter = random.uniform(5, 12)
            else:
                jitter = random.uniform(10, 25)

            info.next_retry_at = now + timedelta(seconds=jitter)
            self.log_warn(f"{info.vat_clean} → THROTTLED ({throttles}) | retry {info.next_retry_at.strftime('%H:%M:%S')}")

            self.root.after(0, self.refresh_trees)
            return info.next_retry_at.timestamp() if info.next_retry_at else None

        elif status in {"TIMEOUT", "ERROR"}:
            info.attempts_hard += 1
            if info.attempts_hard >= self.MAX_ATTEMPTS:
                info.status = "PENDING_MAX"
                info.last_error = result.get("error", "")
                info.next_retry_at = None
                self.log_error(f"{info.vat_clean} → NO VERIFICABLE (máx intentos)")
                self.root.after(0, self.refresh_trees)
                return None
            else:
                info.status = status
                info.last_error = result.get("error", "")
                info.next_retry_at = now + timedelta(seconds=random.uniform(2, 6))
                self.log_warn(f"{info.vat_clean} → {status} ({info.attempts_hard}/{self.MAX_ATTEMPTS})")

                self.root.after(0, self.refresh_trees)
                return info.next_retry_at.timestamp() if info.next_retry_at else None

        else:
            info.attempts_hard += 1
            info.status = "ERROR"
            info.last_error = str(result.get("error", "Error"))
            self.log_error(f"{info.vat_clean} → ERROR")

        self.root.after(0, self.refresh_trees)
        return None

    def validate_vat_with_vies(self, country_code: str, vat_number: str) -> dict:
        # Reutilizamos conexión por hilo para reducir latencia y TIMEOUTs.
        try:
            client = getattr(self._thread_local, "vies_client", None)
            if client is None:
                session = requests.Session()
                transport = Transport(session=session, timeout=self.TIMEOUT)
                client = Client(wsdl=self.VIES_WSDL, transport=transport)
                self._thread_local.vies_client = client

            result = client.service.checkVat(countryCode=country_code, vatNumber=vat_number)

            if result.valid:
                return {
                    "status": "VALID",
                    "vies_name": getattr(result, "name", "") or "",
                    "vies_address": getattr(result, "address", "") or "",
                    "error": "",
                }
            return {"status": "INVALID", "vies_name": "", "vies_address": "", "error": ""}

        except Fault as e:
            msg = str(getattr(e, "message", "")) or str(e)
            detail = str(getattr(e, "detail", ""))
            if "MS_MAX_CONCURRENT_REQ" in msg or "MS_MAX_CONCURRENT_REQ" in detail:
                return {"status": "THROTTLED", "error": "MS_MAX_CONCURRENT_REQ"}
            if "SERVICE_UNAVAILABLE" in msg or "SERVICE_UNAVAILABLE" in detail:
                return {"status": "THROTTLED", "error": "SERVICE_UNAVAILABLE"}
            return {"status": "ERROR", "error": f"SOAP Fault: {msg[:120]}"}

        except TransportError as e:
            # 5xx suele ser temporal
            try:
                status_code = int(getattr(e, "status_code", 0) or 0)
            except Exception:
                status_code = 0
            if status_code in {502, 503, 504}:
                return {"status": "THROTTLED", "error": f"HTTP_{status_code}"}
            return {"status": "TIMEOUT", "error": "TransportError"}

        except (requests.exceptions.Timeout, requests.exceptions.ConnectTimeout):
            return {"status": "TIMEOUT", "error": "TIMEOUT"}

        except requests.exceptions.RequestException as e:
            return {"status": "ERROR", "error": f"Request error: {str(e)[:120]}"}

        except Exception as e:
            return {"status": "ERROR", "error": f"Unexpected: {str(e)[:120]}"}

        finally:
            # La sesión se mantiene viva por hilo; se cerrará al terminar el proceso.
            pass

    def _finish_validation(self) -> None:
        self.processing = False
        self.load_btn.state(["!disabled"])
        self.validate_btn.state(["!disabled"])
        self.retry_btn.state(["!disabled"])

        self.refresh_trees()
        self._update_banner()

        # Calculate summary counts
        pending = sum(1 for v in self.vat_data.values() if v.status not in self.VALIDATED_STATES)
        valid = sum(1 for v in self.vat_data.values() if v.status == "VALID")
        invalid = sum(1 for v in self.vat_data.values() if v.status == "INVALID")
        
        # Log summary
        self.log_info(f"✓ Validación completada: {valid} válidos, {invalid} inválidos, {pending} pendientes")
        
        # Update status bar with summary
        if pending > 0:
            status_msg = (
                f"Validación completada: {valid} válidos, {invalid} inválidos, {pending} pendientes. "
                f"Puedes pulsar 'Reintentar' o 'Validar seleccionado'."
            )
            self.set_status(status_msg, 7000)
        else:
            status_msg = f"Validación completada: {valid} válidos, {invalid} inválidos"
            self.set_status(status_msg, 4500)

    def _update_banner(self) -> None:
        """Update the pending VATs banner with current counts and retry times."""
        pending = sum(1 for v in self.vat_data.values() if v.status not in self.VALIDATED_STATES)
        valid = sum(1 for v in self.vat_data.values() if v.status == "VALID")
        invalid = sum(1 for v in self.vat_data.values() if v.status == "INVALID")
        
        # If validation is complete (no pending), show completion message
        if pending == 0 and (valid > 0 or invalid > 0):
            banner_text = f"✓ Validación terminada: {valid} válidos, {invalid} inválidos"
            self.banner_label.config(text=banner_text)
            self.banner_frame.grid()
            # Auto-hide completion banner after 10 seconds
            self.root.after(10000, lambda: self.banner_frame.grid_remove())
            return
        
        # If there are no VATs at all, hide banner
        if pending == 0:
            self.banner_frame.grid_remove()
            return
        
        # Find next retry time for pending VATs
        now = datetime.now()
        next_retry_time = None
        for v in self.vat_data.values():
            if v.status not in self.VALIDATED_STATES and v.next_retry_at:
                if next_retry_time is None or v.next_retry_at < next_retry_time:
                    next_retry_time = v.next_retry_at
        
        # Update banner text with full summary
        if next_retry_time and next_retry_time > now:
            wait_secs = int((next_retry_time - now).total_seconds())
            banner_text = f"{valid} válidos, {invalid} inválidos, {pending} pendientes (próximo reintento en {wait_secs}s)"
        else:
            banner_text = f"{valid} válidos, {invalid} inválidos, {pending} pendientes"
        
        self.banner_label.config(text=banner_text)
        self.banner_frame.grid()

    def _go_to_pending_tab(self) -> None:
        """Switch to the Pending tab."""
        self.notebook.select(0)

    def undo_last_validated(self) -> None:
        """Undo the last move from Pending to Validated."""
        if not self.undo_stack:
            messagebox.showinfo("Info", "Nada que deshacer.")
            return
        
        if self.processing:
            messagebox.showinfo("Info", "No puedes deshacer mientras hay una validación en curso.")
            return
        
        key, prev_status = self.undo_stack.pop()
        info = self.vat_data.get(key)
        
        if not info:
            return
        
        # Reset to NEW status so it can be validated again with "Validar" batch
        # Clean up all fields that block batch processing
        info.status = "NEW"
        info.last_error = ""
        info.next_retry_at = None
        info.attempts_hard = 0
        info.throttles = 0
        info.last_checked_at = ""
        info.vies_name = ""
        info.vies_address = ""
        info.first_attempt_at = None
        info.auto_retry_count = 0
        
        self.refresh_trees()
        self._update_banner()
        
        # Update undo button state
        if not self.undo_stack:
            self.undo_btn.state(["disabled"])
        
        self.set_status(f"Desecho: {info.vat_clean} (ahora pendiente para validar de nuevo)", 2500)
        self.log_info(f"Desecho: {info.vat_clean} → resetado a Pendiente para revalidar")

    # -------------------------
    # Exit
    # -------------------------

    def exit_app(self) -> None:
        self.on_closing()

    def on_closing(self) -> None:
        msg = "¿Seguro que quieres salir?"
        if self.processing:
            msg = "Hay una validación en curso. ¿Seguro que quieres salir?"
        if messagebox.askyesno("Salir", msg):
            self._stop_event.set()
            self.root.destroy()


def main() -> None:
    root = ttk.Window(themename="flatly")
    app = VATValidatorApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
