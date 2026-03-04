"""VIES VAT Validator (Slango-style)

Aplicación de escritorio para validar números VAT (VIES) a partir de un Excel.
- Carga Excel, detecta columna VAT/NIF automáticamente.
- Separa resultados en pestañas: Pendientes / Validados.
- Maneja limitaciones del servicio (THROTTLED / TIMEOUT) sin bucles infinitos.
- Acción rápida "Abrir VIES": copia el número sin prefijo y abre la web.

Notas:
- VIES es un servicio externo y puede limitar automatizaciones.
- Este programa intenta ser rápido y amable: batches finitos + reintentos manuales.
"""

from __future__ import annotations

import random
import threading
import webbrowser
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, Optional, Tuple, List

import tkinter as tk
from tkinter import filedialog, messagebox

import ttkbootstrap as ttk

from openpyxl import load_workbook, Workbook

from ui_styles import UIStyles
from core.models import (
    VatInfo,
    VatStatus,
    CountryNumber,
    PENDING_STATES,
    VALIDATED_STATES,
    parse_vat,
    get_vat_number_only,
    status_label,
    status_code,
)
from core.validator import ViesValidator
from core.scheduler import ValidationScheduler
from core.callbacks import ValidationCallbacks, BatchSummary


class Tooltip:
    """Simple tooltip helper for ttk widgets."""
    def __init__(self, widget, text: str, delay_ms: int = UIStyles.TOOLTIP_DELAY_MS):
        self.widget = widget
        self.text = text
        self.delay_ms = delay_ms
        self.tipwindow = None
        self.id = None
        
        widget.bind("<Enter>", self.on_enter, add=True)
        widget.bind("<Leave>", self.on_leave, add=True)
    
    def on_enter(self, event):
        self.schedule()
    
    def on_leave(self, event):
        self.unschedule()
        self.hidetip()
    
    def schedule(self):
        self.unschedule()
        self.id = self.widget.after(self.delay_ms, self.showtip)
    
    def unschedule(self):
        if self.id:
            self.widget.after_cancel(self.id)
            self.id = None
    
    def showtip(self):
        if self.tipwindow:
            return
        
        x = self.widget.winfo_rootx() + self.widget.winfo_width() // 2
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 5
        
        self.tipwindow = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        tw.attributes("-topmost", True)
        
        label = tk.Label(tw, text=self.text, background=UIStyles.TOOLTIP_BG, foreground=UIStyles.TOOLTIP_FG,
                         relief=UIStyles.CARD_RELIEF, borderwidth=UIStyles.TOOLTIP_BORDERWIDTH, font=UIStyles.FONT_XSMALL,
                         padx=UIStyles.TOOLTIP_PADX, pady=UIStyles.TOOLTIP_PADY, wraplength=UIStyles.TOOLTIP_WRAPLENGTH)
        label.pack()
    
    def hidetip(self):
        if self.tipwindow:
            self.tipwindow.destroy()
            self.tipwindow = None


class UIThreadCallbacks(ValidationCallbacks):
    """Schedules core callback notifications onto Tk main thread."""

    def __init__(self, app: "VATValidatorApp"):
        self.app = app

    def on_vat_updated(self, key: CountryNumber, vat_info: VatInfo, result: dict) -> None:
        self.app.root.after(0, lambda: self.app._on_vat_updated_main_thread(key, vat_info, result))

    def on_progress(self, done: int, total: int) -> None:
        self.app.root.after(0, lambda: self.app._on_progress_main_thread(done, total))

    def on_banner_update(self, text: str, next_retry_seconds: Optional[int] = None) -> None:
        self.app.root.after(0, lambda: self.app._on_banner_update_main_thread(text, next_retry_seconds))

    def on_batch_finished(self, summary: BatchSummary) -> None:
        self.app.root.after(0, lambda: self.app._on_batch_finished_main_thread(summary))


class VATValidatorApp:
    # Estados from core
    PENDING_STATES = PENDING_STATES
    VALIDATED_STATES = VALIDATED_STATES
    
    # VIES web page
    VIES_WEB = ViesValidator.VIES_WEB

    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("VIES VAT Validator")
        self.root.geometry(f"{UIStyles.WINDOW_WIDTH}x{UIStyles.WINDOW_HEIGHT}")
        self.root.minsize(UIStyles.WINDOW_MIN_WIDTH, UIStyles.WINDOW_MIN_HEIGHT)

        self.selected_file: Optional[Path] = None
        self.processing = False
        self._worker_threads: List[threading.Thread] = []
        self._stop_event = threading.Event()

        # Modelo
        self.vat_data: Dict[CountryNumber, VatInfo] = {}
        self._cache: Dict[CountryNumber, VatInfo] = {}  # cache en memoria por sesión (solo VALID/INVALID)

        # UI state
        self.status_var = tk.StringVar(value="Listo")
        self.log_autoscroll_var = tk.BooleanVar(value=True)

        # Search vars
        self.search_pending_var = tk.StringVar(value="")
        self.search_validated_var = tk.StringVar(value="")

        # Map iids per tree
        self.pending_tree_iids: Dict[CountryNumber, str] = {}
        self.validated_tree_iids: Dict[CountryNumber, str] = {}

        # Stack for undo last validated
        self.undo_stack: List[Tuple[CountryNumber, VatStatus]] = []  # (key, previous_status)

        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

        self.apply_theme()
        self.setup_ui()

        self.log_info("Aplicación iniciada. Carga un Excel para comenzar.")

    # -------------------------
    # THEME / UI
    # -------------------------

    def apply_theme(self) -> None:
        style = ttk.Style()
        
        self.root.configure(bg=UIStyles.BG_MAIN)
        
        # Typography
        style.configure("Treeview", rowheight=UIStyles.TREEVIEW_ROWHEIGHT, background=UIStyles.CARD_BG, fieldbackground=UIStyles.CARD_BG, foreground=UIStyles.TREEVIEW_HEADING_FG)
        style.configure("Treeview.Heading", font=UIStyles.FONT_HEADING, background=UIStyles.TREEVIEW_HEADING_BG, foreground=UIStyles.TREEVIEW_HEADING_FG)
        style.configure("TNotebook", background=UIStyles.CARD_BG)
        style.configure("TNotebook.Tab", padding=(UIStyles.NOTEBOOK_TAB_PADDING_X, UIStyles.NOTEBOOK_TAB_PADDING_Y))

    def setup_ui(self) -> None:
        # Root grid
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        self.root.rowconfigure(1, weight=0)

        # Main container (soft blue background)
        main = tk.Frame(self.root, bg=UIStyles.BG_MAIN)
        main.grid(row=0, column=0, sticky="nsew", padx=UIStyles.MAIN_PADDING, pady=UIStyles.MAIN_PADDING)
        main.columnconfigure(0, weight=1)
        main.rowconfigure(2, weight=1)  # content
        main.rowconfigure(3, weight=0)  # log

        # Header (navy blue background, white text)
        header = tk.Frame(main, bg=UIStyles.HEADER_BG)
        header.grid(row=0, column=0, sticky="ew", pady=(0, UIStyles.CONTENT_PADDING_Y), padx=0)
        header.columnconfigure(0, weight=1)

        title = tk.Label(header, text="VIES VAT Validator", font=UIStyles.FONT_TITLE, bg=UIStyles.HEADER_BG, fg=UIStyles.TEXT_HEADER)
        title.grid(row=0, column=0, sticky="w", padx=UIStyles.CONTENT_PADDING_X, pady=(UIStyles.BUTTON_PADY, 0))
        subtitle = tk.Label(header, text="Validación masiva de números VAT europeos", font=UIStyles.FONT_SUBTITLE, bg=UIStyles.HEADER_BG, fg=UIStyles.TEXT_SUBTITLE)
        subtitle.grid(row=1, column=0, sticky="w", padx=UIStyles.CONTENT_PADDING_X, pady=(2, UIStyles.BUTTON_PADY))

        # Toolbar (simplified)
        toolbar = tk.Frame(main, bg=UIStyles.BG_MAIN)
        toolbar.grid(row=1, column=0, sticky="ew", pady=(UIStyles.BUTTON_PADY, UIStyles.CARD_PADDING_Y), padx=UIStyles.CONTENT_PADDING_X)
        toolbar.columnconfigure(4, weight=1)

        self.load_btn = ttk.Button(toolbar, text="Cargar Excel", bootstyle="primary", padding=(UIStyles.BUTTON_PADX, UIStyles.BUTTON_PADY), command=self.load_excel)
        self.load_btn.grid(row=0, column=0, padx=(0, UIStyles.CARD_PADDING_Y))

        self.validate_btn = ttk.Button(toolbar, text="Validar", bootstyle="primary", padding=(UIStyles.BUTTON_PADX, UIStyles.BUTTON_PADY), command=self.start_validation)
        self.validate_btn.grid(row=0, column=1, padx=(0, UIStyles.CARD_PADDING_Y))
        self.validate_btn.state(["disabled"])

        self.retry_btn = ttk.Button(toolbar, text="Reintentar", bootstyle="primary", padding=(UIStyles.BUTTON_PADX, UIStyles.BUTTON_PADY), command=self.retry_pending)
        self.retry_btn.grid(row=0, column=2, padx=(0, UIStyles.CARD_PADDING_Y))
        self.retry_btn.state(["disabled"])
        Tooltip(self.retry_btn, "Reintenta los pendientes que ya han cumplido el tiempo de espera.")

        self.validate_selected_btn = ttk.Button(toolbar, text="Validar seleccionado", bootstyle="primary", padding=(14, 8), command=self.validate_selected)
        self.validate_selected_btn.grid(row=0, column=3, padx=(0, 8))
        self.validate_selected_btn.state(["disabled"])
        Tooltip(self.validate_selected_btn, "Valida únicamente el VAT seleccionado.")

        self.undo_btn = ttk.Button(toolbar, text="Deshacer último", bootstyle="secondary-outline", padding=(10, 6), command=self.undo_last_validated)
        self.undo_btn.grid(row=0, column=4, padx=(0, 8))
        self.undo_btn.state(["disabled"])

        # Export menu button (right side)
        self.export_menu_btn = ttk.Menubutton(toolbar, text="Exportar ▾", bootstyle="secondary-outline")
        self.export_menu_btn.grid(row=0, column=5, sticky="e", padx=(0, 0))
        self.export_menu_btn.state(["disabled"])
        
        self.export_menu = tk.Menu(self.export_menu_btn, tearoff=0)
        self.export_menu.add_command(label="Exportar todo", command=lambda: self.export_to_excel(scope="all"))
        self.export_menu.add_command(label="Exportar validados", command=lambda: self.export_to_excel(scope="validated"))
        self.export_menu.add_command(label="Exportar pendientes", command=lambda: self.export_to_excel(scope="pending"))
        self.export_menu_btn["menu"] = self.export_menu

        # Content area
        content = tk.Frame(main, bg=UIStyles.BG_MAIN)
        content.grid(row=2, column=0, sticky="nsew")
        content.columnconfigure(0, weight=1)
        content.rowconfigure(0, weight=1)

        # Results card (white frame, no Labelframe)
        results_card = tk.Frame(content, bg=UIStyles.CARD_BG, relief=UIStyles.CARD_RELIEF, borderwidth=UIStyles.CARD_BORDERWIDTH, highlightbackground=UIStyles.CARD_BORDER, highlightthickness=1)
        results_card.grid(row=0, column=0, sticky="nsew", pady=(0, 12), padx=16)
        results_card.columnconfigure(0, weight=1)
        results_card.rowconfigure(2, weight=1)
        
        # Title label on white background
        tk.Label(results_card, text="Resultados", font=UIStyles.FONT_LABEL, bg=UIStyles.CARD_BG, fg=UIStyles.TEXT_PRIMARY, anchor="w", padx=UIStyles.CARD_PADDING_X, pady=UIStyles.CARD_PADDING_Y).grid(row=0, column=0, sticky="ew")

        # Banner for pending VATs (will be shown/hidden dynamically)
        self.banner_frame = tk.Frame(results_card, bg=UIStyles.BANNER_BG, relief=UIStyles.CARD_RELIEF, borderwidth=UIStyles.BANNER_BORDERWIDTH, highlightbackground=UIStyles.BANNER_BORDER, highlightthickness=1)
        self.banner_frame.grid(row=1, column=0, sticky="ew", padx=UIStyles.CARD_PADDING_Y, pady=UIStyles.CARD_PADDING_Y)
        self.banner_frame.columnconfigure(0, weight=1)
        self.banner_frame.grid_remove()  # Hidden by default
        
        # Banner content
        banner_content = tk.Frame(self.banner_frame, bg=UIStyles.BANNER_BG)
        banner_content.grid(row=0, column=0, sticky="ew", padx=UIStyles.BANNER_PADDING_X, pady=UIStyles.BANNER_PADDING_Y)
        banner_content.columnconfigure(0, weight=1)
        
        self.banner_label = tk.Label(banner_content, text="", font=UIStyles.FONT_SMALL, bg=UIStyles.BANNER_BG, fg=UIStyles.BANNER_FG)
        self.banner_label.grid(row=0, column=0, sticky="w")
        
        banner_btn_frame = tk.Frame(banner_content, bg=UIStyles.BANNER_BG)
        banner_btn_frame.grid(row=0, column=1, sticky="e", padx=(UIStyles.BANNER_PADDING_X, 0))
        
        # Custom styled buttons for banner
        self.banner_go_pending_btn = tk.Button(banner_btn_frame, text="Ir a Pendientes", bg=UIStyles.BANNER_BTN_BG, fg=UIStyles.BANNER_BTN_FG, font=UIStyles.FONT_SMALL, relief=UIStyles.CARD_RELIEF, borderwidth=UIStyles.BUTTON_BORDERWIDTH, padx=UIStyles.BUTTON_SMALL_PADX, pady=UIStyles.BUTTON_SMALL_PADY, cursor="hand2", command=self._go_to_pending_tab, activebackground=UIStyles.BANNER_BTN_ACTIVE_BG, activeforeground=UIStyles.BANNER_BTN_ACTIVE_FG)
        self.banner_go_pending_btn.grid(row=0, column=0, padx=(0, UIStyles.BUTTON_SMALL_PADX))
        
        self.banner_retry_btn = tk.Button(banner_btn_frame, text="Reintentar ahora", bg=UIStyles.BANNER_BTN_BG, fg=UIStyles.BANNER_BTN_FG, font=UIStyles.FONT_SMALL, relief=UIStyles.CARD_RELIEF, borderwidth=UIStyles.BUTTON_BORDERWIDTH, padx=UIStyles.BUTTON_SMALL_PADX, pady=UIStyles.BUTTON_SMALL_PADY, cursor="hand2", command=self.retry_pending, activebackground=UIStyles.BANNER_BTN_ACTIVE_BG, activeforeground=UIStyles.BANNER_BTN_ACTIVE_FG)
        self.banner_retry_btn.grid(row=0, column=1)
        Tooltip(self.banner_retry_btn, "Reintenta los VAT que estén listos en este momento.")

        self.notebook = ttk.Notebook(results_card)
        self.notebook.grid(row=2, column=0, sticky="nsew", padx=UIStyles.CARD_PADDING_X, pady=(0, UIStyles.CONTENT_PADDING_Y))

        # Tabs
        self.pending_tab = ttk.Frame(self.notebook, padding=(UIStyles.TREE_PADDING, UIStyles.TREE_PADDING))
        self.validated_tab = ttk.Frame(self.notebook, padding=(UIStyles.TREE_PADDING, UIStyles.TREE_PADDING))
        self.notebook.add(self.pending_tab, text="Pendientes")
        self.notebook.add(self.validated_tab, text="Validados")

        self._build_tab(self.pending_tab, kind="pending")
        self._build_tab(self.validated_tab, kind="validated")

        # Log card (white frame, no Labelframe)
        log_card = tk.Frame(main, bg=UIStyles.CARD_BG, relief=UIStyles.CARD_RELIEF, borderwidth=UIStyles.CARD_BORDERWIDTH, highlightbackground=UIStyles.CARD_BORDER, highlightthickness=1)
        log_card.grid(row=3, column=0, sticky="ew", padx=16, pady=(0, 12))
        log_card.columnconfigure(0, weight=1)
        log_card.rowconfigure(1, weight=1)

        # Header with title and controls
        header_row = tk.Frame(log_card, bg=UIStyles.CARD_BG)
        header_row.grid(row=0, column=0, sticky="ew", padx=12, pady=8)
        header_row.columnconfigure(0, weight=1)
        
        tk.Label(header_row, text="Registro de actividad", font=UIStyles.FONT_LABEL, bg=UIStyles.CARD_BG, fg=UIStyles.TEXT_PRIMARY, anchor="w").grid(row=0, column=0, sticky="w")

        # Controls right
        controls = tk.Frame(header_row, bg=UIStyles.CARD_BG)
        controls.grid(row=0, column=1, sticky="e")

        self.autoscroll_chk = ttk.Checkbutton(controls, text="Auto-scroll", variable=self.log_autoscroll_var, bootstyle="round-toggle")
        self.autoscroll_chk.grid(row=0, column=0, padx=(0, 10))

        ttk.Button(controls, text="Limpiar", bootstyle="secondary-outline", command=self.clear_log).grid(row=0, column=1, padx=(0, 8))
        ttk.Button(controls, text="Copiar", bootstyle="secondary-outline", command=self.copy_log).grid(row=0, column=2)

        # Text + scrollbar (reduced height)
        text_frame = tk.Frame(log_card, bg=UIStyles.CARD_BG)
        text_frame.grid(row=1, column=0, sticky="ew", padx=12, pady=(0, 12))
        text_frame.columnconfigure(0, weight=1)

        self.log_text = tk.Text(
            text_frame,
            height=UIStyles.LOG_HEIGHT,
            wrap=tk.WORD,
            font=UIStyles.FONT_MONOSPACE,
            bg=UIStyles.LOG_BG,
            fg=UIStyles.LOG_FG,
            insertbackground=UIStyles.LOG_FG,
            relief="flat",
            highlightthickness=1,
            highlightbackground=UIStyles.CARD_BORDER,
            highlightcolor="#b6c2cf",
        )
        self.log_text.grid(row=0, column=0, sticky="ew")

        log_scroll = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=self.log_text.yview, bootstyle="round")
        log_scroll.grid(row=0, column=1, sticky="ns")
        self.log_text.configure(yscrollcommand=log_scroll.set)
        self.log_text.configure(state=tk.DISABLED)

        # Tags
        self.log_text.tag_configure("OK", foreground=UIStyles.LOG_OK)
        self.log_text.tag_configure("WARN", foreground=UIStyles.LOG_WARN)
        self.log_text.tag_configure("ERROR", foreground=UIStyles.LOG_ERROR)
        self.log_text.tag_configure("INFO", foreground=UIStyles.LOG_INFO)
        self.log_text.tag_configure("DEBUG", foreground=UIStyles.LOG_DEBUG)

        # Auto-scroll inteligente: si el usuario sube, lo apagamos
        self._install_log_scroll_detection(log_scroll)

        # Footer (status + exit)
        footer = tk.Frame(self.root, bg=UIStyles.BG_MAIN)
        footer.grid(row=1, column=0, sticky="ew", padx=UIStyles.MAIN_PADDING, pady=UIStyles.BUTTON_PADY)
        footer.columnconfigure(0, weight=1)

        self.status_label = tk.Label(footer, textvariable=self.status_var, anchor="w", bg=UIStyles.BG_MAIN, fg=UIStyles.TEXT_PRIMARY, font=UIStyles.FONT_STATUS, padx=UIStyles.CARD_PADDING_X)
        self.status_label.grid(row=0, column=0, sticky="ew")

        self.exit_btn = ttk.Button(footer, text="Salir", bootstyle="danger-outline", width=12, command=self.exit_app)
        self.exit_btn.grid(row=0, column=1, sticky="e", padx=(UIStyles.BUTTON_PADY, 0), pady=(UIStyles.BUTTON_SMALL_PADY, UIStyles.CARD_PADDING_Y))

    def _build_tab(self, parent: ttk.Frame, kind: str) -> None:
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(1, weight=1)

        # Search
        row = ttk.Frame(parent)
        row.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        ttk.Label(row, text="Buscar:").grid(row=0, column=0, sticky="w")

        if kind == "pending":
            entry = ttk.Entry(row, textvariable=self.search_pending_var)
            entry.grid(row=0, column=1, sticky="ew", padx=(8, 0))
            row.columnconfigure(1, weight=1)
            entry.bind("<KeyRelease>", lambda _e: self.refresh_trees())
        else:
            entry = ttk.Entry(row, textvariable=self.search_validated_var)
            entry.grid(row=0, column=1, sticky="ew", padx=(8, 0))
            row.columnconfigure(1, weight=1)
            entry.bind("<KeyRelease>", lambda _e: self.refresh_trees())

        # Tree
        tree_frame = ttk.Frame(parent)
        tree_frame.grid(row=1, column=0, sticky="nsew")
        tree_frame.columnconfigure(0, weight=1)
        tree_frame.rowconfigure(0, weight=1)

        # Define columns based on kind
        if kind == "pending":
            cols = ("VAT", "País", "Número", "Nombre", "Estado", "Intentos", "Última verificación", "Siguiente intento", "Error", "Acción")
        else:
            cols = ("VAT", "País", "Número", "Nombre", "Estado", "Última verificación")

        tree = ttk.Treeview(tree_frame, columns=cols, show="headings", height=UIStyles.TREEVIEW_HEIGHT)
        tree.grid(row=0, column=0, sticky="nsew")

        # Configure headings and columns
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=120, anchor="w")

        # Adjust column widths
        tree.column("VAT", width=UIStyles.COL_VAT)
        tree.column("País", width=UIStyles.COL_COUNTRY, anchor="center")
        tree.column("Número", width=UIStyles.COL_NUMBER)
        tree.column("Nombre", width=UIStyles.COL_NAME)
        tree.column("Estado", width=UIStyles.COL_STATUS)
        
        if kind == "pending":
            tree.column("Intentos", width=UIStyles.COL_ATTEMPTS, anchor="center")
            tree.column("Última verificación", width=UIStyles.COL_LAST_CHECK)
            tree.column("Siguiente intento", width=UIStyles.COL_NEXT_RETRY)
            tree.column("Error", width=UIStyles.COL_ERROR)
            tree.column("Acción", width=UIStyles.COL_ACTION, anchor="center")
        else:
            tree.column("Última verificación", width=UIStyles.COL_LAST_CHECK)

        yscroll = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview, bootstyle="round")
        yscroll.grid(row=0, column=1, sticky="ns")
        tree.configure(yscrollcommand=yscroll.set)

        xscroll = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=tree.xview, bootstyle="round")
        xscroll.grid(row=1, column=0, sticky="ew")
        tree.configure(xscrollcommand=xscroll.set)

        # Zebra striping
        tree.tag_configure("row_even", background=UIStyles.TREE_ROW_EVEN)
        tree.tag_configure("row_odd", background=UIStyles.TREE_ROW_ODD)

        # Row tags (colors)
        tree.tag_configure("VALID", foreground=UIStyles.STATUS_VALID)
        tree.tag_configure("INVALID", foreground=UIStyles.STATUS_INVALID)
        tree.tag_configure("PENDING", foreground=UIStyles.STATUS_PENDING)
        tree.tag_configure("THROTTLED", foreground=UIStyles.STATUS_THROTTLED)
        tree.tag_configure("TIMEOUT", foreground=UIStyles.STATUS_TIMEOUT)
        tree.tag_configure("ERROR", foreground=UIStyles.STATUS_ERROR)
        tree.tag_configure("PENDING_MAX", foreground=UIStyles.STATUS_PENDING_MAX)
        tree.tag_configure("INVALID_FORMAT", foreground=UIStyles.STATUS_INVALID_FORMAT)

        # Bindings
        tree.bind("<<TreeviewSelect>>", self.on_tree_select)
        tree.bind("<Button-1>", self.on_tree_click)
        tree.bind("<Double-1>", self.on_tree_double_click)
        tree.bind("<Motion>", self.on_tree_motion)
        tree.bind("<Button-3>", self.show_context_menu)

        if kind == "pending":
            self.pending_tree = tree
        else:
            self.validated_tree = tree

        # Context menu (minimal)
        # Nota: se construye una sola vez y se reutiliza.
        self.tree_context_menu = tk.Menu(self.root, tearoff=0)
        self.tree_context_menu.add_command(label="Abrir en VIES (web)", command=self.open_vies_web)
        self.tree_context_menu.add_separator()
        self.tree_context_menu.add_command(label="Copiar número VAT", command=lambda: self.copy_vat(number_only=True))
        self.tree_context_menu.add_command(label="Copiar VAT completo", command=lambda: self.copy_vat(number_only=False))

    def _install_log_scroll_detection(self, scrollbar: ttk.Scrollbar) -> None:
        # Detectar cuando el usuario se separa del final: desactivar auto-scroll.
        def on_scroll(*args):
            self.log_text.yview(*args)
            self._check_log_autoscroll_state()

        scrollbar.configure(command=on_scroll)

        def on_mousewheel(event):
            # Windows uses delta, other platforms may differ.
            try:
                self.log_text.yview_scroll(int(-1 * (event.delta / 120)), "units")
            except Exception:
                return
            self._check_log_autoscroll_state()
            return "break"

        self.log_text.bind("<MouseWheel>", on_mousewheel)

    def _check_log_autoscroll_state(self) -> None:
        # yview returns (first, last) in [0,1]
        _, last = self.log_text.yview()
        # If not at bottom, auto-scroll OFF
        if last < 0.999:
            self.log_autoscroll_var.set(False)

    # -------------------------
    # STATUS / LOG
    # -------------------------

    def set_status(self, msg: str, timeout_ms: int = 4000) -> None:
        self.status_var.set(msg)
        self.status_label.update_idletasks()
        if timeout_ms > 0:
            self.root.after(timeout_ms, lambda: self.status_var.set("Listo"))

    def _log(self, level: str, message: str) -> None:
        ts = datetime.now().strftime("%H:%M:%S")
        line = f"[{ts}] {message}\n"
        self.log_text.configure(state=tk.NORMAL)
        self.log_text.insert(tk.END, line, level)
        if self.log_autoscroll_var.get():
            self.log_text.see(tk.END)
        self.log_text.configure(state=tk.DISABLED)

    def log_ok(self, msg: str) -> None:
        self._log("OK", msg)

    def log_warn(self, msg: str) -> None:
        self._log("WARN", msg)

    def log_error(self, msg: str) -> None:
        self._log("ERROR", msg)

    def log_info(self, msg: str) -> None:
        self._log("INFO", msg)

    def log_debug(self, msg: str) -> None:
        self._log("DEBUG", msg)

    def clear_log(self) -> None:
        self.log_text.configure(state=tk.NORMAL)
        self.log_text.delete("1.0", tk.END)
        self.log_text.configure(state=tk.DISABLED)
        self.set_status("Registro limpiado", 2000)

    def copy_log(self) -> None:
        content = self.log_text.get("1.0", tk.END)
        self.root.clipboard_clear()
        self.root.clipboard_append(content)
        self.set_status("Registro copiado al portapapeles", 2500)

    # -------------------------
    # VAT helpers
    # -------------------------

    def accion_text(self, status: VatStatus) -> str:
        if status in {VatStatus.THROTTLED, VatStatus.TIMEOUT, VatStatus.ERROR, VatStatus.PENDING_MAX}:
            return "[[ Abrir VIES ]]"
        return ""

    # -------------------------
    # UI interactions
    # -------------------------

    def _active_tree(self) -> ttk.Treeview:
        # active tab tree: pending or validated
        tab = self.notebook.index(self.notebook.select())
        return self.pending_tree if tab == 0 else self.validated_tree

    def on_tree_select(self, _event=None) -> None:
        tree = self._active_tree()
        sel = tree.selection()
        if sel and not self.processing:
            self.validate_selected_btn.state(["!disabled"])
        else:
            self.validate_selected_btn.state(["disabled"])

    def on_tree_motion(self, event) -> None:
        tree = event.widget
        region = tree.identify_region(event.x, event.y)
        if region != "cell":
            tree.configure(cursor="")
            return
        col = tree.identify_column(event.x)
        row = tree.identify_row(event.y)
        if not row:
            tree.configure(cursor="")
            return
        # Acción is last column
        if col == f"#{len(tree['columns'])}":
            vals = tree.item(row, "values")
            if vals and str(vals[-1]).strip():
                tree.configure(cursor="hand2")
                return
        tree.configure(cursor="")

    def on_tree_click(self, event) -> None:
        tree = event.widget
        region = tree.identify_region(event.x, event.y)
        if region != "cell":
            return
        col = tree.identify_column(event.x)
        row = tree.identify_row(event.y)
        if not row:
            return
        # Acción column
        if col == f"#{len(tree['columns'])}":
            vals = tree.item(row, "values")
            if vals and str(vals[-1]).strip():
                tree.selection_set(row)
                self.open_vies_web()

    def on_tree_double_click(self, event) -> None:
        tree = event.widget
        row = tree.identify_row(event.y)
        if not row:
            return
        vals = tree.item(row, "values")
        if vals and str(vals[-1]).strip():
            tree.selection_set(row)
            self.open_vies_web()

    def show_context_menu(self, event) -> None:
        tree = event.widget
        row = tree.identify_row(event.y)
        if not row:
            return
        tree.selection_set(row)
        self.tree_context_menu.tk_popup(event.x_root, event.y_root)

    def copy_vat(self, number_only: bool) -> None:
        sel = self._get_selected_key()
        if not sel:
            return
        info = self.vat_data.get(sel)
        if not info:
            return
        text = get_vat_number_only(info.vat_clean) if number_only else info.vat_clean
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        if number_only:
            self.set_status(f"Copiado: {text} (País: {info.country})", 3500)
        else:
            self.set_status(f"Copiado: {text}", 2500)

    def open_vies_web(self) -> None:
        sel = self._get_selected_key()
        if not sel:
            return
        info = self.vat_data.get(sel)
        if not info:
            return
        number_only = get_vat_number_only(info.vat_clean)

        def do_ui():
            self.root.clipboard_clear()
            self.root.clipboard_append(number_only)
            
            # Show confirmation dialog
            dialog_msg = (
                f"Se ha copiado al portapapeles el número: {number_only}\n"
                f"País: {info.country}\n\n"
                f"En VIES selecciona el país {info.country} y pega el número.\n\n"
                f"¿Quieres abrir VIES ahora?"
            )
            
            status_msg = f"Copiado: {number_only} (País {info.country}). Confirmación para abrir VIES."
            self.set_status(status_msg, 3500)
            
            # Ask user for confirmation
            if messagebox.askokcancel("Abrir VIES", dialog_msg):
                webbrowser.open(self.VIES_WEB)

        # Execute in UI thread
        self.root.after(0, do_ui)

    # -------------------------
    # Loading / rendering
    # -------------------------

    def refresh_trees(self) -> None:
        # Re-render both trees based on filters
        self._render_pending_tree()
        self._render_validated_tree()

    def _render_pending_tree(self) -> None:
        tree = self.pending_tree
        tree.delete(*tree.get_children())
        self.pending_tree_iids.clear()

        q = self.search_pending_var.get().strip().lower()

        for idx, (key, info) in enumerate(self.vat_data.items()):
            if info.status in self.VALIDATED_STATES:
                continue
            if info.status == VatStatus.INVALID_FORMAT:
                # keep in pending tab
                pass

            if q:
                hay = f"{info.vat_clean} {info.nombre_excel}".lower()
                if q not in hay:
                    continue

            attempts = info.attempts_hard + info.throttles
            last_checked = info.last_checked_at
            retry = ""
            if info.next_retry_at:
                retry = info.next_retry_at.strftime("%H:%M:%S")

            values = (
                info.vat_clean,
                info.country,
                info.number,
                info.nombre_excel,
                status_label(info.status),
                attempts,
                last_checked,
                retry,
                info.last_error,
                self.accion_text(info.status),
            )
            iid = str(key)
            row_tag = "row_even" if idx % 2 == 0 else "row_odd"
            tree.insert("", "end", iid=iid, values=values, tags=(row_tag, status_code(info.status)))
            self.pending_tree_iids[key] = iid

    def _render_validated_tree(self) -> None:
        tree = self.validated_tree
        tree.delete(*tree.get_children())
        self.validated_tree_iids.clear()

        q = self.search_validated_var.get().strip().lower()

        for idx, (key, info) in enumerate(self.vat_data.items()):
            if info.status not in self.VALIDATED_STATES:
                continue
            if q:
                hay = f"{info.vat_clean} {info.nombre_excel} {info.vies_name}".lower()
                if q not in hay:
                    continue

            last_checked = info.last_checked_at
            # Only 6 columns for validated tree
            values = (
                info.vat_clean,
                info.country,
                info.number,
                info.nombre_excel or info.vies_name,
                status_label(info.status),
                last_checked,
            )
            iid = str(key)
            row_tag = "row_even" if idx % 2 == 0 else "row_odd"
            tree.insert("", "end", iid=iid, values=values, tags=(row_tag, status_code(info.status)))
            self.validated_tree_iids[key] = iid

    def _get_selected_key(self) -> Optional[CountryNumber]:
        tree = self._active_tree()
        sel = tree.selection()
        if not sel:
            return None
        iid = sel[0]
        try:
            # iid is str(key) where key is tuple; recover by searching
            vals = tree.item(iid, "values")
            if not vals:
                return None
            country = vals[1]
            number = vals[2]
            return (country, number)
        except Exception:
            return None

    # -------------------------
    # Excel load / export
    # -------------------------

    def load_excel(self) -> None:
        if self.processing:
            messagebox.showinfo("Info", "Hay una validación en curso.")
            return

        file_path = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
        )
        if not file_path:
            return

        self.selected_file = Path(file_path)
        self.vat_data.clear()

        # UI state
        self.retry_btn.state(["disabled"])
        self.export_menu_btn.state(["disabled"])
        self.validate_selected_btn.state(["disabled"])

        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active

            # Detect header row: find columns containing NIF/VAT
            header_row = None
            nif_col = None
            name_col = None

            for r in range(1, min(30, ws.max_row) + 1):
                row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
                row_text = " ".join([str(v).strip().upper() for v in row_vals if v is not None])
                if any(k in row_text for k in ["NIF", "VAT", "VAT NUMBER", "VAT_NUMBER"]):
                    header_row = r
                    # Find column indices
                    for c in range(1, ws.max_column + 1):
                        val = ws.cell(row=r, column=c).value
                        if val is None:
                            continue
                        t = str(val).strip().upper()
                        if t in {"NIF", "VAT", "VAT NUMBER", "VAT_NUMBER"}:
                            nif_col = c
                        if t in {"NOMBRE", "NAME", "RAZON SOCIAL", "RAZÓN SOCIAL"}:
                            name_col = c
                    break

            if header_row is None or nif_col is None:
                # fallback: if first column looks like VAT values
                header_row = 1
                nif_col = 1

            # Load data
            for r in range(header_row + 1, ws.max_row + 1):
                raw_vat = ws.cell(row=r, column=nif_col).value
                if raw_vat is None:
                    continue

                country, number, vat_clean = parse_vat(raw_vat)
                nombre_excel = ""
                if name_col:
                    v = ws.cell(row=r, column=name_col).value
                    if v is not None:
                        nombre_excel = str(v).strip()

                if not vat_clean:
                    continue

                if not country or not number:
                    # invalid format
                    # Use placeholder country/number to keep stable key: try to split anyway
                    # We'll store under ("", vat_clean)
                    key = ("", vat_clean)
                    self.vat_data[key] = VatInfo(vat_clean=vat_clean, country="", number="", nombre_excel=nombre_excel, status=VatStatus.INVALID_FORMAT, last_error="Formato inválido")
                    continue

                key = (country, number)

                # Cache hit (VALID/INVALID): reuse
                cached = self._cache.get(key)
                if cached and cached.status in self.VALIDATED_STATES:
                    info = VatInfo(
                        vat_clean=vat_clean,
                        country=country,
                        number=number,
                        nombre_excel=nombre_excel or cached.nombre_excel,
                        status=cached.status,
                        vies_name=cached.vies_name,
                        vies_address=cached.vies_address,
                        attempts_hard=cached.attempts_hard,
                        throttles=cached.throttles,
                        last_checked_at=cached.last_checked_at,
                        last_error=cached.last_error,
                    )
                    self.vat_data[key] = info
                    continue

                self.vat_data[key] = VatInfo(vat_clean=vat_clean, country=country, number=number, nombre_excel=nombre_excel)

            self.refresh_trees()

            self.log_ok(f"Cargados {len(self.vat_data)} VATs únicos desde {self.selected_file.name}")
            self.set_status(f"Excel cargado: {len(self.vat_data)} VATs. Pulsa 'Validar' para comenzar.", 3500)

            # Enable controls
            self.validate_btn.state(["!disabled"])
            self.export_menu_btn.state(["!disabled"])

        except Exception as e:
            self.log_error(f"Error al cargar Excel: {e}")
            messagebox.showerror("Error", f"No se pudo cargar el archivo: {e}")

    def export_to_excel(self, scope: str = "all") -> None:
        if not self.vat_data:
            messagebox.showinfo("Info", "No hay datos para exportar")
            return

        if scope not in {"all", "pending", "validated"}:
            scope = "all"

        initial_name = "vat_results.xlsx"
        if self.selected_file:
            suffix = {"all": "validated", "pending": "pending", "validated": "validated"}[scope]
            initial_name = f"{self.selected_file.stem}_{suffix}.xlsx"

        out = filedialog.asksaveasfilename(
            title="Guardar resultados",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("All", "*.*")],
            initialfile=initial_name,
        )
        if not out:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "VAT Results"

            headers = [
                "VAT_CLEAN",
                "COUNTRY",
                "NUMBER",
                "NOMBRE_EXCEL",
                "STATUS",
                "ATTEMPTS",
                "THROTTLES",
                "LAST_CHECKED_AT",
                "NEXT_RETRY_AT",
                "VIES_NAME",
                "VIES_ADDRESS",
                "ERROR",
            ]
            for i, h in enumerate(headers, start=1):
                ws.cell(row=1, column=i, value=h)

            row = 2
            for key, info in self.vat_data.items():
                if scope == "pending" and info.status in self.VALIDATED_STATES:
                    continue
                if scope == "validated" and info.status not in self.VALIDATED_STATES:
                    continue

                attempts = info.attempts_hard + info.throttles
                ws.cell(row=row, column=1, value=info.vat_clean)
                ws.cell(row=row, column=2, value=info.country)
                ws.cell(row=row, column=3, value=info.number)
                ws.cell(row=row, column=4, value=info.nombre_excel)
                ws.cell(row=row, column=5, value=status_code(info.status))
                ws.cell(row=row, column=6, value=attempts)
                ws.cell(row=row, column=7, value=info.throttles)
                ws.cell(row=row, column=8, value=info.last_checked_at)
                ws.cell(row=row, column=9, value=info.next_retry_at.strftime("%Y-%m-%d %H:%M:%S") if info.next_retry_at else "")
                ws.cell(row=row, column=10, value=info.vies_name)
                ws.cell(row=row, column=11, value=info.vies_address)
                ws.cell(row=row, column=12, value=info.last_error)
                row += 1

            wb.save(out)
            self.log_ok(f"Exportado: {Path(out).name}")
            self.set_status(f"Exportado: {Path(out).name}", 3500)
            messagebox.showinfo("Éxito", f"Resultados exportados a:\n{out}")
        except Exception as e:
            self.log_error(f"Error al exportar: {e}")
            messagebox.showerror("Error", f"No se pudo exportar: {e}")

    # -------------------------
    # Validation flow
    # -------------------------

    def start_validation(self) -> None:
        if self.processing:
            return

        to_validate = [(k, v) for k, v in self.vat_data.items() if v.status == VatStatus.NEW and v.country and v.number]
        if not to_validate:
            return

        self.processing = True
        self._stop_event.clear()
        self.load_btn.state(["disabled"])
        self.validate_btn.state(["disabled"])
        self.retry_btn.state(["disabled"])
        self.validate_selected_btn.state(["disabled"])

        self.log_info("============================================================")
        self.log_info(f"Iniciando validación de {len(to_validate)} VATs...")

        t = threading.Thread(target=self._validate_batch_worker, args=(to_validate,), daemon=True)
        self._worker_threads = [t]
        t.start()

    def retry_pending(self) -> None:
        if self.processing:
            return

        now = datetime.now()
        to_retry = []
        next_retry_time = None

        for k, v in self.vat_data.items():
            if v.is_retryable():
                if v.next_retry_at is None or v.next_retry_at <= now:
                    to_retry.append((k, v))
                elif next_retry_time is None or v.next_retry_at < next_retry_time:
                    next_retry_time = v.next_retry_at

        if not to_retry:
            if next_retry_time:
                wait_secs = int((next_retry_time - now).total_seconds())
                msg = f"Aún no se puede reintentar. Próximo reintento recomendado en {wait_secs}s."
                messagebox.showinfo("Reintentar", msg)
            else:
                messagebox.showinfo("Info", "No hay VATs pendientes listos para reintentar.")
            return

        self.processing = True
        self._stop_event.clear()
        self.load_btn.state(["disabled"])
        self.retry_btn.state(["disabled"])
        self.validate_selected_btn.state(["disabled"])

        self.log_info("============================================================")
        self.log_info(f"Reintentando {len(to_retry)} VATs pendientes...")

        t = threading.Thread(target=self._validate_batch_worker, args=(to_retry,), daemon=True)
        self._worker_threads = [t]
        t.start()

    def validate_selected(self) -> None:
        if self.processing:
            return
        key = self._get_selected_key()
        if not key:
            messagebox.showinfo("Info", "Selecciona un VAT.")
            return

        info = self.vat_data.get(key)
        if not info or not info.country or not info.number:
            return

        # 1 intento inmediato
        self.processing = True
        self._stop_event.clear()
        self.load_btn.state(["disabled"])
        self.validate_btn.state(["disabled"])
        self.retry_btn.state(["disabled"])
        self.validate_selected_btn.state(["disabled"])

        self.log_info(f"Validando seleccionado: {info.vat_clean}")

        t = threading.Thread(target=self._validate_batch_worker, args=([(key, info)],), daemon=True)
        self._worker_threads = [t]
        t.start()

    def _validate_batch_worker(self, items: List[Tuple[CountryNumber, VatInfo]]) -> None:
        """Worker that validates a batch of VATs using the scheduler."""
        callbacks = UIThreadCallbacks(self)

        # Create scheduler and run validation
        scheduler = ValidationScheduler(self.vat_data, callbacks, self._stop_event)
        scheduler.validate_batch(items)

    def _on_vat_updated_main_thread(self, key: CountryNumber, info: VatInfo, result: dict) -> None:
        self._apply_result(info, result)

        # Cache VALID/INVALID
        if info.status in self.VALIDATED_STATES:
            self._cache[key] = info

    def _on_progress_main_thread(self, done: int, total: int) -> None:
        self.set_status(f"Validando… {done}/{total}", 0)

    def _on_banner_update_main_thread(self, text: str, next_retry_seconds: Optional[int] = None) -> None:
        if text:
            self.banner_label.config(text=text)
            self.banner_frame.grid()
        else:
            self._update_banner()

    def _on_batch_finished_main_thread(self, summary: BatchSummary) -> None:
        self._finish_validation(summary)

    def _apply_result(self, info: VatInfo, result: dict) -> Optional[float]:
        """Apply validation result to VatInfo and update UI."""
        status: VatStatus = result.get("status")
        now = datetime.now()
        info.last_checked_at = now.strftime("%Y-%m-%d %H:%M:%S")
        prev_status = info.status  # Save previous status for undo

        if status == VatStatus.VALID:
            info.status = VatStatus.VALID
            info.vies_name = result.get("vies_name", "")
            info.vies_address = result.get("vies_address", "")
            info.last_error = ""
            info.next_retry_at = None
            self.log_ok(f"{info.vat_clean} → VALID")
            # Save to undo stack if transitioning to validated
            if prev_status not in self.VALIDATED_STATES:
                self.undo_stack.append(((info.country, info.number), prev_status))
                self.root.after(0, lambda: self.undo_btn.state(["!disabled"]))

            self.root.after(0, self.refresh_trees)
            return None

        elif status == VatStatus.INVALID:
            info.status = VatStatus.INVALID
            info.vies_name = ""
            info.vies_address = ""
            info.last_error = ""
            info.next_retry_at = None
            self.log_warn(f"{info.vat_clean} → INVALID")
            # Save to undo stack if transitioning to validated
            if prev_status not in self.VALIDATED_STATES:
                self.undo_stack.append(((info.country, info.number), prev_status))
                self.root.after(0, lambda: self.undo_btn.state(["!disabled"]))

            self.root.after(0, self.refresh_trees)
            return None

        elif status == VatStatus.THROTTLED:
            info.throttles += 1
            info.status = VatStatus.THROTTLED
            info.last_error = result.get("error", "MS_MAX_CONCURRENT_REQ")

            # Retry suggestion
            throttles = info.throttles
            if throttles == 1:
                jitter = random.uniform(2, 7)
            elif throttles == 2:
                jitter = random.uniform(5, 12)
            else:
                jitter = random.uniform(10, 25)

            info.next_retry_at = now + timedelta(seconds=jitter)
            self.log_warn(f"{info.vat_clean} → THROTTLED ({throttles}) | retry {info.next_retry_at.strftime('%H:%M:%S')}")

            self.root.after(0, self.refresh_trees)
            return info.next_retry_at.timestamp() if info.next_retry_at else None

        elif status in {VatStatus.TIMEOUT, VatStatus.ERROR}:
            info.attempts_hard += 1
            max_attempts = ValidationScheduler.MAX_ATTEMPTS
            if info.attempts_hard >= max_attempts:
                info.status = VatStatus.PENDING_MAX
                info.last_error = result.get("error", "")
                info.next_retry_at = None
                self.log_error(f"{info.vat_clean} → NO VERIFICABLE (máx intentos)")
                self.root.after(0, self.refresh_trees)
                return None
            else:
                info.status = status
                info.last_error = result.get("error", "")
                info.next_retry_at = now + timedelta(seconds=random.uniform(2, 6))
                self.log_warn(f"{info.vat_clean} → {status_code(status)} ({info.attempts_hard}/{max_attempts})")

                self.root.after(0, self.refresh_trees)
                return info.next_retry_at.timestamp() if info.next_retry_at else None

        else:
            info.attempts_hard += 1
            info.status = VatStatus.ERROR
            info.last_error = str(result.get("error", "Error"))
            self.log_error(f"{info.vat_clean} → ERROR")

        self.root.after(0, self.refresh_trees)
        return None

    def _finish_validation(self, summary: Optional[BatchSummary] = None) -> None:
        self.processing = False
        self.load_btn.state(["!disabled"])
        self.validate_btn.state(["!disabled"])
        self.retry_btn.state(["!disabled"])

        self.refresh_trees()
        self._update_banner()

        # Calculate summary counts
        if summary is not None:
            pending = summary.pending
            valid = summary.valid
            invalid = summary.invalid
        else:
            pending = sum(1 for v in self.vat_data.values() if v.status not in self.VALIDATED_STATES)
            valid = sum(1 for v in self.vat_data.values() if v.status == VatStatus.VALID)
            invalid = sum(1 for v in self.vat_data.values() if v.status == VatStatus.INVALID)
        
        # Log summary
        self.log_info(f"✓ Validación completada: {valid} válidos, {invalid} inválidos, {pending} pendientes")
        
        # Update status bar with summary
        if pending > 0:
            status_msg = (
                f"Validación completada: {valid} válidos, {invalid} inválidos, {pending} pendientes. "
                f"Puedes pulsar 'Reintentar' o 'Validar seleccionado'."
            )
            self.set_status(status_msg, 7000)
        else:
            status_msg = f"Validación completada: {valid} válidos, {invalid} inválidos"
            self.set_status(status_msg, 4500)

    def _update_banner(self) -> None:
        """Update the pending VATs banner with current counts and retry times."""
        pending = sum(1 for v in self.vat_data.values() if v.status not in self.VALIDATED_STATES)
        valid = sum(1 for v in self.vat_data.values() if v.status == VatStatus.VALID)
        invalid = sum(1 for v in self.vat_data.values() if v.status == VatStatus.INVALID)
        
        # If validation is complete (no pending), show completion message
        if pending == 0 and (valid > 0 or invalid > 0):
            banner_text = f"✓ Validación terminada: {valid} válidos, {invalid} inválidos"
            self.banner_label.config(text=banner_text)
            self.banner_frame.grid()
            # Auto-hide completion banner after 10 seconds
            self.root.after(10000, lambda: self.banner_frame.grid_remove())
            return
        
        # If there are no VATs at all, hide banner
        if pending == 0:
            self.banner_frame.grid_remove()
            return
        
        # Find next retry time for pending VATs
        now = datetime.now()
        next_retry_time = None
        for v in self.vat_data.values():
            if v.status not in self.VALIDATED_STATES and v.next_retry_at:
                if next_retry_time is None or v.next_retry_at < next_retry_time:
                    next_retry_time = v.next_retry_at
        
        # Update banner text with full summary
        if next_retry_time and next_retry_time > now:
            wait_secs = int((next_retry_time - now).total_seconds())
            banner_text = f"{valid} válidos, {invalid} inválidos, {pending} pendientes (próximo reintento en {wait_secs}s)"
        else:
            banner_text = f"{valid} válidos, {invalid} inválidos, {pending} pendientes"
        
        self.banner_label.config(text=banner_text)
        self.banner_frame.grid()

    def _go_to_pending_tab(self) -> None:
        """Switch to the Pending tab."""
        self.notebook.select(0)

    def undo_last_validated(self) -> None:
        """Undo the last move from Pending to Validated."""
        if not self.undo_stack:
            messagebox.showinfo("Info", "Nada que deshacer.")
            return
        
        if self.processing:
            messagebox.showinfo("Info", "No puedes deshacer mientras hay una validación en curso.")
            return
        
        key, prev_status = self.undo_stack.pop()
        info = self.vat_data.get(key)
        
        if not info:
            return
        
        # Reset to NEW status so it can be validated again with "Validar" batch
        # Clean up all fields that block batch processing
        info.status = VatStatus.NEW
        info.last_error = ""
        info.next_retry_at = None
        info.attempts_hard = 0
        info.throttles = 0
        info.last_checked_at = ""
        info.vies_name = ""
        info.vies_address = ""
        info.first_attempt_at = None
        info.auto_retry_count = 0
        
        self.refresh_trees()
        self._update_banner()
        
        # Update undo button state
        if not self.undo_stack:
            self.undo_btn.state(["disabled"])
        
        self.set_status(f"Desecho: {info.vat_clean} (ahora pendiente para validar de nuevo)", 2500)
        self.log_info(f"Desecho: {info.vat_clean} → resetado a Pendiente para revalidar")

    # -------------------------
    # Exit
    # -------------------------

    def exit_app(self) -> None:
        self.on_closing()

    def on_closing(self) -> None:
        msg = "¿Seguro que quieres salir?"
        if self.processing:
            msg = "Hay una validación en curso. ¿Seguro que quieres salir?"
        if messagebox.askyesno("Salir", msg):
            self._stop_event.set()
            self.root.destroy()


def main() -> None:
    root = ttk.Window(themename="flatly")
    app = VATValidatorApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
