"""Manejo de lectura y escritura de archivos Excel con datos VAT."""

from pathlib import Path
from typing import Dict, Optional, Tuple, List
from openpyxl import load_workbook, Workbook

from .models import VatInfo, VatStatus, CountryNumber, parse_vat, status_code


def detect_vat_column(ws, max_rows: int = 30) -> Tuple[Optional[int], Optional[int]]:
    """Detecta automáticamente las columnas de VAT y Nombre en una hoja Excel.
    
    Busca en las primeras filas headers que contengan palabras clave como NIF, VAT, NOMBRE, etc.
    
    Args:
        ws: Worksheet de openpyxl
        max_rows: Número máximo de filas a inspeccionar para encontrar headers
        
    Returns:
        Tupla (header_row, nif_column, name_column) o (1, 1, None) si no encuentra
    """
    header_row = None
    nif_col = None
    name_col = None

    for r in range(1, min(max_rows, ws.max_row) + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        row_text = " ".join([str(v).strip().upper() for v in row_vals if v is not None])
        if any(k in row_text for k in ["NIF", "VAT", "VAT NUMBER", "VAT_NUMBER"]):
            header_row = r
            # Encuentra columnas específicas
            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=r, column=c).value
                if val is None:
                    continue
                t = str(val).strip().upper()
                if t in {"NIF", "VAT", "VAT NUMBER", "VAT_NUMBER"}:
                    nif_col = c
                if t in {"NOMBRE", "NAME", "RAZON SOCIAL", "RAZÓN SOCIAL"}:
                    name_col = c
            break

    # Fallback si no encuentra headers
    if header_row is None:
        header_row = 1
    if nif_col is None:
        nif_col = 1

    return header_row, nif_col, name_col


def load_excel(file_path: Path) -> Dict[CountryNumber, VatInfo]:
    """Carga datos VAT desde archivo Excel.
    
    Lee un Excel, detecta automáticamente columnas VAT y Nombre, y carga en memoria.
    En caso de VAT con formato inválido, los almacena bajo clave ("", vat_clean).
    
    Args:
        file_path: Ruta al archivo .xlsx or .xls
        
    Returns:
        Diccionario con clave (país, número) → VatInfo
        
    Raises:
        Exception: Si hay problema al abrir o leer el Excel
    """
    vat_data: Dict[CountryNumber, VatInfo] = {}
    
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    header_row, nif_col, name_col = detect_vat_column(ws)

    # Carga datos
    for r in range(header_row + 1, ws.max_row + 1):
        raw_vat = ws.cell(row=r, column=nif_col).value
        if raw_vat is None:
            continue

        country, number, vat_clean = parse_vat(raw_vat)
        nombre_excel = ""
        if name_col:
            v = ws.cell(row=r, column=name_col).value
            if v is not None:
                nombre_excel = str(v).strip()

        if not vat_clean:
            continue

        if not country or not number:
            # Formato inválido: guardar bajo clave especial
            key = ("", vat_clean)
            vat_data[key] = VatInfo(
                vat_clean=vat_clean,
                country="",
                number="",
                nombre_excel=nombre_excel,
                status=VatStatus.INVALID_FORMAT,
                last_error="Formato inválido"
            )
            continue

        key = (country, number)
        vat_data[key] = VatInfo(
            vat_clean=vat_clean,
            country=country,
            number=number,
            nombre_excel=nombre_excel
        )

    return vat_data


def save_excel(file_path: Path, vat_data: Dict[CountryNumber, VatInfo], scope: str = "all") -> None:
    """Exporta datos VAT a archivo Excel.
    
    Crea un Excel con todos los campos de VatInfo y estados de validación.
    
    Args:
        file_path: Ruta destino para guardar .xlsx
        vat_data: Diccionario de VatInfo a exportar
        scope: "all", "pending" o "validated"
        
    Raises:
        Exception: Si hay problema al escribir Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "VAT Results"

    headers = [
        "VAT_CLEAN",
        "COUNTRY",
        "NUMBER",
        "NOMBRE_EXCEL",
        "STATUS",
        "ATTEMPTS",
        "THROTTLES",
        "LAST_CHECKED_AT",
        "NEXT_RETRY_AT",
        "VIES_NAME",
        "VIES_ADDRESS",
        "ERROR",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)

    row = 2
    for key, info in vat_data.items():
        # Filtrado por scope
        if scope == "pending":
            if info.status in {VatStatus.VALID, VatStatus.INVALID}:
                continue
        elif scope == "validated":
            if info.status not in {VatStatus.VALID, VatStatus.INVALID}:
                continue

        attempts = info.attempts_hard + info.throttles
        ws.cell(row=row, column=1, value=info.vat_clean)
        ws.cell(row=row, column=2, value=info.country)
        ws.cell(row=row, column=3, value=info.number)
        ws.cell(row=row, column=4, value=info.nombre_excel)
        ws.cell(row=row, column=5, value=status_code(info.status))
        ws.cell(row=row, column=6, value=attempts)
        ws.cell(row=row, column=7, value=info.throttles)
        ws.cell(row=row, column=8, value=info.last_checked_at)
        ws.cell(row=row, column=9, value=info.next_retry_at.strftime("%Y-%m-%d %H:%M:%S") if info.next_retry_at else "")
        ws.cell(row=row, column=10, value=info.vies_name)
        ws.cell(row=row, column=11, value=info.vies_address)
        ws.cell(row=row, column=12, value=info.last_error)
        row += 1

    wb.save(file_path)
